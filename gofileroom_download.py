"""
Script tự động hóa GOFILEROOM để tìm kiếm client và export document
Hỗ trợ cả single download và multiple download theo từng page
"""

import csv
import os
import re
import time
import sys
import logging
import shutil
import zipfile
import traceback
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import config
from document_mapping import get_document_category, get_all_categories
from file_handler import (
    rename_file_with_doc_id, move_file, move_csv_to_storage,
    move_zip_to_storage, extract_zip, remove_file, find_file_in_zip_folder
)
from excel_handler import ExcelHandler
from email_handler import create_email_handler_from_config
from models import Document, Client, BASE_DOWNLOAD_DIR
from utils import resource_path, load_env_config


# Thiết lập logging
log_file_name = 'gofileroom_download.log'
log_file_path = resource_path(log_file_name)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path, encoding='utf-8'),
        # logging.FileHandler('gofileroom_download.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class GofileRoomDownloader:
    """Class chính để xử lý download documents từ GOFILEROOM"""
    
    def __init__(self):
        """Khởi tạo downloader"""
        self.driver = None
        self.wait = None
        self.workbook = None
        self.client_list_sheet = None
        self.document_list_sheet = None
        
        # Load config từ .env
        self.config = self._load_config()
        
        # Cấu hình download - đọc từ .env hoặc dùng giá trị mặc định
        self.download_dir = BASE_DOWNLOAD_DIR
        self.csv_dir = os.path.join(self.download_dir, "0_csv_")
        self.zip_dir = os.path.join(self.download_dir, "0_zip_")
        
        # Tạo các thư mục cần thiết
        os.makedirs(self.download_dir, exist_ok=True)
        os.makedirs(self.csv_dir, exist_ok=True)
        os.makedirs(self.zip_dir, exist_ok=True)
        
        # Biến tracking
        self.current_client_info = {}
        self.total_documents = 0
        self.downloaded_documents = 0
        self.error_description = ""
        
        # Khởi tạo ExcelHandler (sẽ được set sau khi load workbook)
        self.excel_handler = None
        
        # Khởi tạo EmailHandler
        self.email_handler = create_email_handler_from_config(self.config)
        
        # Excel file properties
        self.excel_file_name = self.config.get('CLIENT_LIST_FILE_NAME', 'download_gofileroom_data.xlsx')
        self.excel_file_path = resource_path(self.excel_file_name)
        
        logger.info("GofileRoomDownloader initialized")
    
    def _load_config(self):
        """Đọc cấu hình từ file .env"""
        return load_env_config(raise_on_not_found=True)
    
    def login(self):
        """Thực hiện login vào GOFILEROOM"""
        try:
            logger.info("Đang thực hiện login vào GOFILEROOM...")
            
            # Đọc username và password từ config
            username = self.config.get('USERNAME')
            password = self.config.get('PASSWORD')
            
            if not username:
                error_msg = "Không tìm thấy USERNAME trong file .env"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Chuyển về ngữ cảnh mặc định
            self.driver.switch_to.default_content()
            
            # Tìm và nhập username
            try:
                # Thử tìm bằng ID trước
                username_input = self.wait.until(
                    EC.presence_of_element_located((By.ID, "txtLogin"))
                )
            except TimeoutException:
                # Nếu không tìm thấy bằng ID, thử xpath
                username_input = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/form/div[1]/div/div/input"))
                )
            
            # Xóa nội dung cũ và nhập username
            username_input.clear()
            username_input.send_keys(username)
            logger.info(f"Đã nhập username: {username}")
            time.sleep(1)
            
            # Tìm và click button Sign In
            try:
                # Thử tìm bằng ID trước
                sign_in_btn = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "btnSignIn1"))
                )
            except TimeoutException:
                # Nếu không tìm thấy bằng ID, thử xpath
                sign_in_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div[2]/form/button"))
                )
            
            # Kiểm tra button có bị disable không
            btn_class = sign_in_btn.get_attribute('class') or ''
            if 'disabled' in btn_class or not sign_in_btn.is_enabled():
                error_msg = "Button Sign In bị disable"
                logger.error(error_msg)
                raise Exception(error_msg)
            
            sign_in_btn.click()
            logger.info("Đã click button Sign In")
            
            # Chờ 15s để tự chuyển hướng sang trang BASE_URL
            logger.info("Đang chờ chuyển hướng sau khi login (15 giây)...")
            time.sleep(10)
            
            # Kiểm tra URL sau khi login
            current_url = self.driver.current_url
            logger.info(f"URL sau khi login: {current_url}")
            
            if config.LOGIN_URL in current_url:
                error_msg = f"Vẫn còn ở trang login sau khi thực hiện login: {current_url}"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            if config.BASE_URL in current_url:
                logger.info("Đã login thành công và chuyển về trang chính")
                return True
            else:
                logger.warning(f"URL sau khi login ({current_url}) khác với BASE_URL ({config.BASE_URL})")
                logger.warning("Tiếp tục xử lý...")
                return True
            
        except Exception as e:
            error_msg = f"Lỗi khi thực hiện login: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False
    
    def setup_driver(self):
        """Thiết lập Chrome WebDriver - kết nối với Chrome đang chạy sẵn"""
        try:
            logger.info("Đang kết nối với Chrome đang chạy...")
            
            chrome_options = Options()
            prefs = {
                "download.default_directory": self.download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            }
            chrome_options.add_experimental_option("prefs", prefs)
            chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            logger.info("Đã kết nối với Chrome đang chạy thành công")
            
            self.driver.implicitly_wait(config.IMPLICIT_WAIT)
            self.driver.set_page_load_timeout(config.PAGE_LOAD_TIMEOUT)
            self.wait = WebDriverWait(self.driver, config.EXPLICIT_WAIT)
            
        except Exception as e:
            logger.error(f"Lỗi khi kết nối với Chrome: {str(e)}")
            logger.error(traceback.format_exc())
            logger.info("Hướng dẫn:")
            logger.info("1. Đóng tất cả Chrome đang chạy")
            logger.info("2. Chạy lệnh: chrome.exe --remote-debugging-port=9222")
            logger.info("3. Đăng nhập vào GOFILEROOM trên Chrome vừa mở")
            logger.info("4. Chạy lại script này")
            raise
    
    def read_client_list(self, excel_file_path):
        """Đọc danh sách client từ file Excel"""
        try:
            logger.info(f"Đang đọc file Excel: {excel_file_path}")
            
            if not os.path.exists(excel_file_path):
                raise FileNotFoundError(f"File Excel không tồn tại: {excel_file_path}")
            
            self.workbook = load_workbook(excel_file_path)
            client_list_sheet_name = self.config.get('CLIENT_LIST_SHEET_NAME', 'Client List GFR')
            document_list_sheet_name = self.config.get('DOCUMENT_LIST_SHEET_NAME', 'Download Document Log')
            
            # Lấy các sheet
            if client_list_sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Sheet '{client_list_sheet_name}' không tồn tại trong file Excel")
            if document_list_sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Sheet '{document_list_sheet_name}' không tồn tại trong file Excel")
            
            self.client_list_sheet = self.workbook[client_list_sheet_name]
            self.document_list_sheet = self.workbook[document_list_sheet_name]
            
            # Khởi tạo ExcelHandler
            self.excel_handler = ExcelHandler(
                self.workbook, self.client_list_sheet, self.document_list_sheet, excel_file_path
            )
            
            # Đọc header để xác định vị trí cột
            header_row = [cell.value for cell in self.client_list_sheet[1]]
            logger.info(f"Header row: {header_row}")
            
            # Tìm index các cột
            try:
                status_idx = header_row.index('Status')
                description_idx = header_row.index('Description')
                client_name_idx = header_row.index('Client Name')
                client_number_idx = header_row.index('Client Number')
                client_email_idx = header_row.index('Client Email')
                total_docs_idx = header_row.index('Total Documents')
                num_files_downloaded_idx = header_row.index('Number Of Files Downloaded')
                client_folder_path_idx = header_row.index('Client Folder Path')
            except ValueError as e:
                logger.error(f"Không tìm thấy cột cần thiết trong header: {str(e)}")
                raise
            
            # Đọc danh sách client có Status = "Pending"
            client_data = []
            for row_idx, row in enumerate(self.client_list_sheet.iter_rows(min_row=2, values_only=False), start=2):
                status_cell = row[status_idx]
                if status_cell.value and str(status_cell.value).strip() == "Pending":
                    client_name_cell = row[client_name_idx]
                    client_number_cell = row[client_number_idx]
                    
                    if client_name_cell.value and client_number_cell.value:
                        client_data.append({
                            'row_index': row_idx,
                            'status_cell': row[status_idx],
                            'description_cell': row[description_idx],
                            'client_name_cell': row[client_name_idx],
                            'client_number_cell': row[client_number_idx],
                            'client_email_cell': row[client_email_idx],
                            'total_documents_cell': row[total_docs_idx],
                            'num_files_downloaded_cell': row[num_files_downloaded_idx],
                            'client_folder_path_cell': row[client_folder_path_idx],
                            'client_name': str(client_name_cell.value).strip(),
                            'client_number': str(client_number_cell.value).strip(),
                        })
            
            logger.info(f"Đã đọc {len(client_data)} client có Status = Pending")
            return client_data
            
        except Exception as e:
            logger.error(f"Lỗi khi đọc file Excel: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def _clean_download_dir(self):
        """Xóa tất cả file trong download_dir, giữ lại các folder"""
        try:
            logger.info("Đang dọn sạch download_dir (chỉ xóa file, giữ folder)...")
            files_deleted = 0
            
            for item in os.listdir(self.download_dir):
                item_path = os.path.join(self.download_dir, item)
                try:
                    if os.path.isfile(item_path):
                        os.remove(item_path)
                        files_deleted += 1
                except Exception as e:
                    logger.warning(f"Không thể xóa file {item_path}: {str(e)}")
            
            logger.info(f"Đã xóa {files_deleted} file trong download_dir")
            return True
            
        except Exception as e:
            logger.error(f"Lỗi khi dọn sạch download_dir: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def _count_files_in_download_dir(self):
        """Đếm số file (không phải folder) trong download_dir"""
        try:
            count = 0
            for item in os.listdir(self.download_dir):
                item_path = os.path.join(self.download_dir, item)
                if os.path.isfile(item_path):
                    count += 1
            return count
        except Exception as e:
            logger.error(f"Lỗi khi đếm file trong download_dir: {str(e)}")
            return 0
    
    def search_client(self, client_name):
        """Tìm kiếm client theo tên"""
        try:
            logger.info(f"Đang tìm kiếm client: {client_name}")
            
            # Chuyển về ngữ cảnh mặc định
            self.driver.switch_to.default_content()
            
            # Tìm và chuyển vào iframe
            try:
                iframe = self.wait.until(
                    EC.presence_of_element_located(config.SEARCH_CLIENT_IFRAME_LOCATOR)
                )
                self.driver.switch_to.frame(iframe)
            except TimeoutException:
                error_msg = f"Timeout: Không tìm thấy iframe với locator: {config.SEARCH_CLIENT_IFRAME_LOCATOR}"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            # Chuyển vào frame level 2
            FRAME_LEVEL_2 = (By.NAME, "mainFrame")
            try:
                self.wait.until(
                    EC.frame_to_be_available_and_switch_to_it(FRAME_LEVEL_2)
                )
            except TimeoutException:
                error_msg = f"Timeout: Không tìm thấy frame level 2: {FRAME_LEVEL_2}"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            # Tìm ô tìm kiếm
            search_input = None
            try:
                search_input = self.wait.until(
                    EC.element_to_be_clickable(config.SEARCH_INPUT_LOCATOR)
                )
            except TimeoutException:
                logger.warning("Không tìm thấy với selector chính, thử các selector khác...")
                for selector in config.SEARCH_INPUT_ALTERNATIVES:
                    try:
                        search_input = self.driver.find_element(selector[0], selector[1])
                        if search_input.is_displayed() and search_input.is_enabled():
                            break
                        search_input = None
                    except NoSuchElementException:
                        continue
                
                if not search_input:
                    error_msg = "Không tìm thấy ô tìm kiếm với bất kỳ selector nào"
                    logger.error(error_msg)
                    self.error_description = error_msg
                    return False
            
            # Kích hoạt ô tìm kiếm nếu cần
            if not search_input.is_enabled():
                try:
                    search_input.click()
                    time.sleep(1)
                except:
                    self.driver.execute_script("arguments[0].click();", search_input)
                    time.sleep(1)
            
            # Xóa nội dung cũ và nhập tên client
            try:
                search_input.clear()
            except:
                search_input.send_keys(Keys.CONTROL + "a")
                search_input.send_keys(Keys.DELETE)
            
            search_input.send_keys(client_name)
            search_input.send_keys(Keys.RETURN)
            
            logger.info(f"Đã nhập tên client '{client_name}' và ấn Enter")
            return True
            
        except Exception as e:
            error_msg = f"Lỗi khi tìm kiếm client '{client_name}': {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False
    
    def check_client_exists(self, client_name, client_number):
        """Kiểm tra client có tồn tại không và lấy số lượng documents"""
        try:
            logger.info("Đang chờ kết quả tìm kiếm client...")
            
            # Chờ phần tử gốc của cây thư mục client
            client_tree_root = self.wait.until(
                EC.presence_of_element_located(config.CLIENT_TREE_ROOT_LOCATOR)
            )
            time.sleep(1)
            
            try:
                client_list_ul = client_tree_root.find_element(By.XPATH, "./ul")
                client_item_a_tags = client_list_ul.find_elements(By.TAG_NAME, "a")
                
                if not client_item_a_tags:
                    error_msg = "Không tìm thấy client items trong cây thư mục"
                    logger.warning(error_msg)
                    self.error_description = error_msg
                    return False, None
                
                # Tìm client phù hợp
                client_item = None
                for a_tag in client_item_a_tags:
                    a_tag_text = str(a_tag.text)
                    search_text = f"{client_name} | {client_number}".lower()
                    if a_tag_text.lower().startswith(search_text):
                        client_item = a_tag
                        break
                
                if not client_item:
                    error_msg = f"Không tìm thấy client '{client_name} | {client_number}' trong cây thư mục"
                    logger.warning(error_msg)
                    self.error_description = error_msg
                    return False, None
                
                # Click vào client để load document list
                client_item.click()
                logger.info("Đã click vào client để load Document List")
                time.sleep(2)
                
                # Lấy số lượng documents từ text (format: "Client Name | Number (count)")
                pattern = r'\((\d+)\)'
                match = re.search(pattern, client_item.text)
                if match:
                    number_documents = int(match.group(1))
                    logger.info(f"Tìm thấy client với {number_documents} documents")
                    return True, number_documents
                else:
                    logger.warning("Không tìm thấy số lượng documents trong text")
                    return True, 0
                
            except NoSuchElementException:
                error_msg = "Không tìm thấy thẻ <ul> bên trong Client Tree"
                logger.warning(error_msg)
                self.error_description = error_msg
                return False, None
                
        except TimeoutException:
            error_msg = "Timeout: Không tìm thấy Client Tree Root"
            logger.error(error_msg)
            self.error_description = error_msg
            return False, None
        except Exception as e:
            error_msg = f"Lỗi khi kiểm tra client tồn tại: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False, None
    
    def _wait_for_file_download(self, expected_extension="", timeout=120):
        """Chờ file được tải về trong download_dir"""
        try:
            logger.info(f"Đang chờ file được tải về (timeout: {timeout}s)...")
            start_time = time.time()
            
            while time.time() - start_time < timeout:
                # Kiểm tra file tạm
                temp_files = [f for f in os.listdir(self.download_dir) 
                             if f.endswith(('.crdownload', '.tmp'))]
                if temp_files:
                    time.sleep(2)
                    continue
                
                # Lấy danh sách file
                if expected_extension:
                    files = [f for f in os.listdir(self.download_dir)
                            if f.endswith(expected_extension) and 
                            os.path.isfile(os.path.join(self.download_dir, f))]
                else:
                    files = [f for f in os.listdir(self.download_dir)
                            if os.path.isfile(os.path.join(self.download_dir, f))]
                
                if files:
                    # Lấy file mới nhất
                    full_paths = [os.path.join(self.download_dir, f) for f in files]
                    latest_file = max(full_paths, key=os.path.getmtime)
                    
                    # Kiểm tra file đã hoàn thành chưa
                    if not os.path.exists(latest_file + ".crdownload"):
                        logger.info(f"File đã tải xong: {os.path.basename(latest_file)}")
                        return True, latest_file
                
                time.sleep(2)
            
            error_msg = f"Timeout: Không tìm thấy file sau {timeout} giây"
            logger.error(error_msg)
            self.error_description = error_msg
            return False, None
            
        except Exception as e:
            error_msg = f"Lỗi khi chờ file tải về: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False, None
    
    def download_csv_list(self, client_name, client_number):
        """Tải CSV file chứa danh sách documents"""
        try:
            logger.info("Đang tải CSV file danh sách documents...")
            
            # Click nút Export List
            EXPORT_LIST_BTN_LOCATOR = (By.XPATH, "//button[contains(text(), 'Export List')]")
            btn_export_list = self.wait.until(
                EC.element_to_be_clickable(EXPORT_LIST_BTN_LOCATOR)
            )
            btn_export_list.click()
            logger.info("Đã click nút Export List")
            
            # Chờ file CSV được tải về
            success, csv_file_path = self._wait_for_file_download(expected_extension=".csv", timeout=120)
            if not success:
                return False, None
            
            # Kiểm tra số lượng file trong download_dir
            file_count = self._count_files_in_download_dir()
            if file_count > 1:
                error_msg = f"Phát hiện {file_count} file trong download_dir, chỉ mong đợi 1 file CSV"
                logger.warning(error_msg)
                self.error_description = error_msg
                return False, None
            
            # Move file CSV vào thư mục 0_csv_
            csv_filename = os.path.basename(csv_file_path)
            csv_dest_path = os.path.join(self.csv_dir, csv_filename)
            
            # Nếu file đã tồn tại, đổi tên
            if os.path.exists(csv_dest_path):
                base, ext = os.path.splitext(csv_filename)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                csv_filename = f"{base}_{timestamp}{ext}"
                csv_dest_path = os.path.join(self.csv_dir, csv_filename)
            
            shutil.move(csv_file_path, csv_dest_path)
            logger.info(f"Đã move CSV file vào: {csv_dest_path}")
            
            return True, csv_dest_path
            
        except Exception as e:
            error_msg = f"Lỗi khi tải CSV file: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False, None
    
    def read_csv_file(self, csv_file_path, client_name_check, client_number_check):
        """Đọc file CSV và trả về dictionary với Document ID làm key"""
        try:
            logger.info(f"Đang đọc CSV file: {csv_file_path}")
            
            if not os.path.exists(csv_file_path):
                error_msg = f"CSV file không tồn tại: {csv_file_path}"
                logger.error(error_msg)
                self.error_description = error_msg
                return None
            
            documents_data = {}
            
            with open(csv_file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                
                # Đọc header
                try:
                    header = next(reader)
                    header = [h.strip() for h in header]
                    
                    # Tìm index các cột
                    doc_id_index = header.index("Document ID")
                    client_name_index = header.index("Client Name")
                    client_number_index = header.index("Client Number")
                    
                except (StopIteration, ValueError) as e:
                    error_msg = f"Lỗi khi đọc header CSV: {str(e)}"
                    logger.error(error_msg)
                    self.error_description = error_msg
                    return None
                
                # Đọc dữ liệu
                try:
                    first_data_row = next(reader)
                except StopIteration:
                    logger.warning("CSV file chỉ có header, không có dữ liệu")
                    return documents_data
                
                # Kiểm tra Client Name/Number từ hàng đầu tiên
                actual_client_name = first_data_row[client_name_index].strip().strip('"')
                actual_client_number = first_data_row[client_number_index].strip().strip('"')
                
                if actual_client_number != client_number_check:
                    error_msg = (f"CSV file không khớp: "
                               f"Expected '{client_name_check}'/'{client_number_check}', "
                               f"Got '{actual_client_name}'/'{actual_client_number}'")
                    logger.error(error_msg)
                    self.error_description = error_msg
                    return None
                
                # Đọc hàng đầu tiên
                doc_id = first_data_row[doc_id_index].strip().strip('"')
                documents_data[doc_id] = {
                    header[i]: first_data_row[i].strip().strip('"') for i in range(len(header))
                }
                
                # Đọc các hàng còn lại
                for row in reader:
                    if not row:
                        continue
                    doc_id = row[doc_id_index].strip().strip('"')
                    documents_data[doc_id] = {
                        header[i]: row[i].strip().strip('"') for i in range(len(header))
                    }
                
                # Tính expected_download_file_name cho mỗi document
                for doc_id, doc_info in documents_data.items():
                    expected_name_items = []
                    if doc_info.get("Client Name"):
                        expected_name_items.append(str(doc_info["Client Name"]))
                    if doc_info.get("Year"):
                        expected_name_items.append(str(doc_info["Year"]))
                    if doc_info.get("Document Type"):
                        expected_name_items.append(str(doc_info["Document Type"]))
                    if doc_info.get("Description"):
                        expected_name_items.append(str(doc_info["Description"]))
                    
                    expected_name = "_".join(expected_name_items)
                    expected_name = re.sub(r'[\\/:*?"<>|]', '', expected_name)
                    file_type = doc_info.get("File Type", "pdf")
                    expected_name = f"{expected_name}.{file_type}"
                    doc_info["expected_download_file_name"] = expected_name
            
            logger.info(f"Đã đọc {len(documents_data)} documents từ CSV")
            return documents_data
            
        except Exception as e:
            error_msg = f"Lỗi khi đọc CSV file: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return None
    
    def _find_document_row_in_excel(self, client_name, client_number, document_id, year, file_type):
        """Tìm row trong document_list_sheet với thông tin đã cho"""
        try:
            # Đọc header để tìm index các cột
            header_row = [cell.value for cell in self.document_list_sheet[1]]
            
            try:
                client_name_idx = header_row.index('Client Name')
                client_number_idx = header_row.index('Client Number')
                doc_id_idx = header_row.index('Document ID')
                year_idx = header_row.index('Year')
                file_type_idx = header_row.index('File Type')
            except ValueError as e:
                logger.error(f"Không tìm thấy cột trong header: {str(e)}")
                return None
            
            # Tìm row phù hợp
            for row_idx, row in enumerate(self.document_list_sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    row_client_name = str(row[client_name_idx].value or "").strip()
                    row_client_number = str(row[client_number_idx].value or "").strip()
                    row_doc_id = str(row[doc_id_idx].value or "").strip()
                    row_year = str(row[year_idx].value or "").strip()
                    row_file_type = str(row[file_type_idx].value or "").strip()
                    
                    if (row_client_name == client_name and
                        row_client_number == client_number and
                        row_doc_id == document_id and
                        row_year == str(year or "") and
                        row_file_type == file_type):
                        return row_idx
                except:
                    continue
            
            return None
            
        except Exception as e:
            logger.error(f"Lỗi khi tìm document row trong Excel: {str(e)}")
            logger.error(traceback.format_exc())
            return None
    
    def log_documents_to_excel(self, csv_documents_dict, client_name, client_number):
        """Log danh sách documents vào Excel sheet, trả về dict mapping doc_id -> row_index"""
        try:
            logger.info("Đang log documents vào Excel...")
            
            # Đọc header để tìm index các cột
            header_row = [cell.value for cell in self.document_list_sheet[1]]
            
            try:
                download_status_idx = header_row.index('Download Status')
                download_desc_idx = header_row.index('Download Description')
                client_name_idx = header_row.index('Client Name')
                client_number_idx = header_row.index('Client Number')
                file_name_idx = header_row.index('File Name')
                file_path_idx = header_row.index('File Path')
                file_section_idx = header_row.index('File Section')
                doc_type_idx = header_row.index('Document Type')
                description_idx = header_row.index('Description')
                year_idx = header_row.index('Year')
                doc_date_idx = header_row.index('Document Date')
                file_size_idx = header_row.index('File Size')
                doc_id_idx = header_row.index('Document ID')
                file_type_idx = header_row.index('File Type')
                download_time_idx = header_row.index('Download time')
            except ValueError as e:
                logger.error(f"Không tìm thấy cột trong header: {str(e)}")
                return {}
            
            doc_id_to_row_index = {}
            new_rows_count = 0
            existing_rows_count = 0
            
            for doc_id, doc_info in csv_documents_dict.items():
                # Kiểm tra xem document đã tồn tại chưa
                row_index = self._find_document_row_in_excel(
                    client_name, client_number, doc_id,
                    doc_info.get("Year", ""), doc_info.get("File Type", "")
                )
                
                if row_index:
                    # Document đã tồn tại, chỉ lưu row_index
                    doc_id_to_row_index[doc_id] = row_index
                    existing_rows_count += 1
                else:
                    # Document chưa tồn tại, thêm mới
                    new_row = [
                        "",  # Download Status
                        "",  # Download Description
                        client_name,  # Client Name
                        client_number,  # Client Number
                        doc_info.get("expected_download_file_name", ""),  # File Name
                        "",  # File Path
                        doc_info.get("File Section", ""),  # File Section
                        doc_info.get("Document Type", ""),  # Document Type
                        doc_info.get("Description", ""),  # Description
                        doc_info.get("Year", ""),  # Year
                        doc_info.get("Document Date", ""),  # Document Date
                        doc_info.get("File Size", ""),  # File Size
                        doc_id,  # Document ID
                        doc_info.get("File Type", ""),  # File Type
                        ""  # Download time
                    ]
                    self.document_list_sheet.append(new_row)
                    # Lấy row index vừa thêm
                    new_row_index = self.document_list_sheet.max_row
                    doc_id_to_row_index[doc_id] = new_row_index
                    new_rows_count += 1
            
            logger.info(f"Đã log documents: {new_rows_count} mới, {existing_rows_count} đã tồn tại")
            return doc_id_to_row_index
            
        except Exception as e:
            error_msg = f"Lỗi khi log documents vào Excel: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return {}
    
    def _sanitize_folder_name(self, folder_name):
        """
        Validate và format tên folder để đảm bảo an toàn cho hệ thống file
        
        Args:
            folder_name (str): Tên folder gốc
            
        Returns:
            str: Tên folder đã được sanitize
        """
        if not folder_name:
            return ""
        
        sanitized = (str(folder_name)
                    .replace(" | ", "_")
                    .replace(" ", "_")
                    .replace("(", "")
                    .replace(")", "")
                    .replace('"', "")
                    .strip())
        
        return sanitized
    
    def _create_folder(self, folder_name, parent_folder):
        """
        Tạo folder trong parent folder
        
        Args:
            folder_name (str): Tên folder cần tạo (đã được sanitize)
            parent_folder (str): Đường dẫn thư mục cha
            
        Returns:
            tuple: (success: bool, full_path: str)
                - success: True nếu tạo thành công, False nếu có lỗi
                - full_path: Đường dẫn đầy đủ đến folder nếu thành công, "" nếu thất bại
        """
        try:
            if not folder_name:
                logger.error("Tên folder không được rỗng")
                return False, ""
            
            if not parent_folder:
                logger.error("Parent folder không được rỗng")
                return False, ""
            
            full_path = os.path.join(parent_folder, folder_name)
            
            if not os.path.exists(full_path):
                os.makedirs(full_path, exist_ok=True)
                logger.debug(f"Đã tạo folder: {full_path}")
            
            return True, full_path
            
        except Exception as e:
            logger.error(f"Lỗi khi tạo folder '{folder_name}' trong '{parent_folder}': {str(e)}")
            logger.error(traceback.format_exc())
            return False, ""
    
    def _initialize_category_folders(self, client_folder_path):
        """
        Khởi tạo các category folders trong client folder nếu chưa có
        
        Args:
            client_folder_path (str): Đường dẫn đến client folder
            
        Returns:
            bool: True nếu thành công, False nếu có lỗi
        """
        try:
            categories = get_all_categories()
            
            for category in categories:
                category_status, category_path = self._create_folder(category, client_folder_path)
                if not category_status:
                    logger.warning(f"Không thể tạo category folder '{category}' trong '{client_folder_path}'")
                    # Tiếp tục tạo các folder khác, không dừng lại
            
            logger.debug(f"Đã khởi tạo category folders trong: {client_folder_path}")
            return True
            
        except Exception as e:
            logger.error(f"Lỗi khi khởi tạo category folders: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def _get_safe_client_dir(self, dir_str, root_folder, is_client_folder=False):
        """
        Helper method để sanitize tên folder và tạo folder
        (Giữ lại để backward compatibility, sử dụng _sanitize_folder_name và _create_folder bên trong)
        
        Args:
            dir_str (str): Tên folder gốc
            root_folder (str): Đường dẫn thư mục cha
            is_client_folder (bool): True nếu đây là client folder, sẽ khởi tạo category folders
            
        Returns:
            tuple: (success: bool, full_path: str)
        """
        sanitized_name = self._sanitize_folder_name(dir_str)
        success, full_path = self._create_folder(sanitized_name, root_folder)
        
        # Nếu là client folder và tạo thành công, khởi tạo category folders
        if success and is_client_folder:
            self._initialize_category_folders(full_path)
        
        return success, full_path
    
    def update_document_status_in_excel(self, doc_id, row_index, download_status, download_description, 
                                        file_name="", file_path="", download_time=""):
        """Cập nhật status của document trong Excel"""
        try:
            if not row_index:
                return False
            
            # Đọc header để tìm index các cột
            header_row = [cell.value for cell in self.document_list_sheet[1]]
            
            try:
                download_status_idx = header_row.index('Download Status')
                download_desc_idx = header_row.index('Download Description')
                file_name_idx = header_row.index('File Name')
                file_path_idx = header_row.index('File Path')
                download_time_idx = header_row.index('Download time')
            except ValueError as e:
                logger.error(f"Không tìm thấy cột trong header: {str(e)}")
                return False
            
            # Cập nhật các cell
            row = self.document_list_sheet[row_index]
            if download_status:
                row[download_status_idx].value = download_status
            if download_description:
                row[download_desc_idx].value = download_description
            if file_name:
                row[file_name_idx].value = file_name
            if file_path:
                row[file_path_idx].value = file_path
            if download_time:
                row[download_time_idx].value = download_time
            
            return True
            
        except Exception as e:
            logger.error(f"Lỗi khi cập nhật document status trong Excel: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def click_export_single_file(self, row_index, document):
        """
        Click export button cho một file riêng lẻ, tải file, đổi tên, move vào đúng folder
        
        Args:
            row_index (int): Index của row trong document table (0-based)
            document (Document): Document object
            
        Returns:
            tuple: (success: bool, file_path: str, error_msg: str)
        """
        try:
            logger.info(f"Đang export file cho document {document.document_id} (row {row_index})")
            
            # Kiểm tra file đã tồn tại chưa
            if document.file_exists:
                logger.info(f"File đã tồn tại: {document.document_file_path}")
                return True, document.document_file_path, ""
            
            # Tìm document row trong table
            document_table = self.wait.until(
                EC.presence_of_element_located(config.DOCUMENT_TABLE_LOCATOR)
            )
            document_rows = document_table.find_elements(
                config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                config.DOCUMENT_TABLE_DIV_LOCATOR[1]
            )
            
            if row_index >= len(document_rows):
                error_msg = f"Row index {row_index} vượt quá số lượng rows ({len(document_rows)})"
                logger.error(error_msg)
                return False, "", error_msg
            
            row = document_rows[row_index]
            
            # Kiểm tra document ID trong row
            document_data_cells = row.find_elements(
                config.DOCUMENT_DATA_CELL_LOCATOR[0], 
                config.DOCUMENT_DATA_CELL_LOCATOR[1]
            )
            
            if len(document_data_cells) < 10:
                error_msg = "Không tìm thấy đủ cells trong document row"
                logger.error(error_msg)
                return False, "", error_msg
            
            row_doc_id = document_data_cells[9].text.strip()
            if row_doc_id != document.document_id:
                error_msg = f"Document ID không khớp: expected {document.document_id}, got {row_doc_id}"
                logger.error(error_msg)
                return False, "", error_msg
            
            # Tìm và click export button
            document_first_cell = row.find_elements(
                config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[0], 
                config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[1]
            )
            if not document_first_cell:
                error_msg = "Không tìm thấy first cell của document row"
                logger.error(error_msg)
                return False, "", error_msg
            
            btns = document_first_cell[0].find_elements(By.TAG_NAME, "button")
            if len(btns) < 3:
                error_msg = "Không tìm thấy export button"
                logger.error(error_msg)
                return False, "", error_msg
            
            btn_export = btns[2]
            
            # Kiểm tra button có bị disable không
            if not btn_export.is_enabled():
                error_msg = f"Export button bị disable cho document {document.document_id}"
                logger.warning(error_msg)
                return False, "", error_msg
            
            btn_export.click()
            logger.info(f"Đã click export button cho document {document.document_id}")
            time.sleep(5)
            
            # Chờ file được tải về
            success, downloaded_file_path = self._wait_for_file_download(timeout=120)
            if not success:
                error_msg = f"Không tìm thấy file sau khi download cho document {document.document_id}"
                logger.error(error_msg)
                return False, "", error_msg
            
            # Kiểm tra số lượng file
            file_count = self._count_files_in_download_dir()
            if file_count > 1:
                error_msg = f"Phát hiện {file_count} file trong download_dir, chỉ mong đợi 1 file"
                logger.warning(error_msg)
                # Xóa file để tiếp tục
                try:
                    os.remove(downloaded_file_path)
                except:
                    pass
                return False, "", error_msg
            
            # Kiểm tra tên file (tùy chọn, có thể bỏ qua nếu không khớp)
            downloaded_file_name = os.path.basename(downloaded_file_path)
            base_name, ext = os.path.splitext(downloaded_file_name)
            expected_base, expected_ext = os.path.splitext(document.document_name_without_id)
            
            if base_name != expected_base or ext != expected_ext:
                logger.warning(
                    f"Tên file không khớp cho document {document.document_id}: "
                    f"expected '{document.document_name_without_id}', got '{downloaded_file_name}'"
                )
                # Vẫn tiếp tục xử lý
            
            # Đổi tên file để thêm document_id
            success, renamed_file_path, error_msg = rename_file_with_doc_id(
                downloaded_file_path, document.document_id
            )
            if not success:
                logger.error(error_msg)
                return False, "", error_msg
            
            # Move file vào đúng folder
            if not document.document_folder_path:
                error_msg = f"Document folder path không được set cho document {document.document_id}"
                logger.error(error_msg)
                return False, "", error_msg
            
            # Đảm bảo folder tồn tại
            if not os.path.exists(document.document_folder_path):
                os.makedirs(document.document_folder_path, exist_ok=True)
            
            final_dest = os.path.join(document.document_folder_path, document.document_name_with_id)
            success, error_msg = move_file(renamed_file_path, final_dest)
            if not success:
                logger.error(error_msg)
                return False, "", error_msg
            
            logger.info(f"Đã move file vào: {final_dest}")
            return True, final_dest, ""
            
        except Exception as e:
            error_msg = f"Lỗi khi export single file: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            return False, "", error_msg
    
    def export_single(self, client_name, client_number, csv_documents_dict, doc_id_to_row_index, client_target_dir):
        """Export single document (khi total_documents = 1)"""
        try:
            logger.info(f"Đang export single document cho client: {client_name}")
            
            # Lấy document duy nhất
            if len(csv_documents_dict) != 1:
                error_msg = f"export_single được gọi nhưng có {len(csv_documents_dict)} documents"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            doc_id = list(csv_documents_dict.keys())[0]
            doc_info = csv_documents_dict[doc_id]
            expected_file_name = doc_info.get("expected_download_file_name", "")
            doc_year = doc_info.get("Year", "").strip()
            
            # Xác định category folder để kiểm tra file đã tồn tại
            file_section = doc_info.get("File Section", "").strip()
            document_type = doc_info.get("Document Type", "").strip()
            description = doc_info.get("Description", "").strip()
            category_folder = get_document_category(file_section, document_type, description)
            
            # get category folder để kiểm tra
            category_dir = os.path.join(client_target_dir, category_folder)
            
            # Kiểm tra file đã tồn tại chưa
            expected_base, expected_ext = os.path.splitext(expected_file_name)
            if doc_year:
                year_dir_status, year_dir = self._create_folder(doc_year, category_dir)
                if not year_dir_status:
                    error_msg = f"Không thể tạo year folder: {doc_year}"
                    logger.error(error_msg)
                    row_index = doc_id_to_row_index.get(doc_id)
                    if row_index:
                        self.update_document_status_in_excel(
                            doc_id, row_index, "Error", error_msg
                        )
                    return False
                file_path = os.path.join(category_dir, doc_year, f"{expected_base}_{doc_id}{expected_ext}")
            else:
                file_path = os.path.join(category_dir, f"{expected_base}_{doc_id}{expected_ext}")
            
            if os.path.exists(file_path):
                logger.info(f"File đã tồn tại: {file_path}")
                self.downloaded_documents = 1
                download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Success", "File already exists",
                        expected_file_name, file_path, download_time
                    )
                return True
            
            # Tìm document row trong table và click export
            document_table = self.wait.until(
                EC.presence_of_element_located(config.DOCUMENT_TABLE_LOCATOR)
            )
            document_rows = document_table.find_elements(
                config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                config.DOCUMENT_TABLE_DIV_LOCATOR[1]
            )
            
            if len(document_rows) <= 0:
                error_msg = "Không tìm thấy document row trong table"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            row = document_rows[0]  # Row đầu tiên sau header
            document_data_cells = row.find_elements(
                config.DOCUMENT_DATA_CELL_LOCATOR[0], 
                config.DOCUMENT_DATA_CELL_LOCATOR[1]
            )
            
            if len(document_data_cells) < 10:
                error_msg = "Không tìm thấy đủ cells trong document row"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            row_doc_id = document_data_cells[9].text.strip()
            if row_doc_id != doc_id:
                error_msg = f"Document ID không khớp: expected {doc_id}, got {row_doc_id}"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            # Click export button
            document_first_cell = row.find_elements(
                config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[0], 
                config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[1]
            )
            if not document_first_cell:
                error_msg = "Không tìm thấy first cell của document row"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            btns = document_first_cell[0].find_elements(By.TAG_NAME, "button")
            if len(btns) < 3:
                error_msg = "Không tìm thấy export button"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            btn_export = btns[2]
            btn_export.click()
            logger.info("Đã click export button")
            time.sleep(5)
            
            # Chờ file được tải về
            success, downloaded_file_path = self._wait_for_file_download(timeout=120)
            if not success:
                error_msg = "Không tìm thấy file sau khi download"
                logger.error(error_msg)
                self.error_description = error_msg
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Error", error_msg
                    )
                return False
            
            # Kiểm tra số lượng file
            file_count = self._count_files_in_download_dir()
            if file_count > 1:
                error_msg = f"Phát hiện {file_count} file trong download_dir, chỉ mong đợi 1 file"
                logger.warning(error_msg)
                self.error_description = error_msg
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Warning", error_msg
                    )
                return False
            
            # Kiểm tra tên file
            downloaded_file_name = os.path.basename(downloaded_file_path)
            base_name, ext = os.path.splitext(downloaded_file_name)
            expected_base, expected_ext = os.path.splitext(expected_file_name)
            
            if base_name != expected_base or ext != expected_ext:
                error_msg = (f"Tên file không khớp: expected '{expected_file_name}', "
                           f"got '{downloaded_file_name}'")
                logger.error(error_msg)
                self.error_description = error_msg
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Error", error_msg
                    )
                return False
            
            # Đổi tên file để thêm document_id
            new_file_name = f"{base_name}_{doc_id}{ext}"
            renamed_file_path = os.path.join(category_dir, new_file_name)
            try:
                os.rename(downloaded_file_path, renamed_file_path)
            except Exception as e:
                error_msg = f"Lỗi khi đổi tên file: {str(e)}"
                logger.error(error_msg)
                self.error_description = error_msg
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Error", error_msg
                    )
                return False
            
            if doc_year:
                final_dest = os.path.join(category_dir, doc_year, new_file_name)
            else:
                final_dest = os.path.join(category_dir, new_file_name)
            
            # Move file vào category hoặc category/year folder
            try:
                shutil.move(renamed_file_path, final_dest)
                logger.info(f"Đã move file vào: {final_dest}")
                self.downloaded_documents = 1
                
                # Cập nhật Excel
                download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Success", "File downloaded successfully",
                        new_file_name, final_dest, download_time
                    )
                
                return True
                
            except Exception as e:
                error_msg = f"Lỗi khi move file: {str(e)}"
                logger.error(error_msg)
                logger.error(traceback.format_exc())
                self.error_description = error_msg
                row_index = doc_id_to_row_index.get(doc_id)
                if row_index:
                    self.update_document_status_in_excel(
                        doc_id, row_index, "Error", error_msg
                    )
                return False
            
        except Exception as e:
            error_msg = f"Lỗi tổng quát khi export single: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False
    
    def export_page_with_multiple_button(self, client_object, page_num, total_pages):
        """
        Export một page bằng cách click export button (export multiple)
        Nếu không click được thì sẽ chuyển sang export từng file lẻ
        
        Args:
            client_object (Client): Client object
            page_num (int): Số page hiện tại
            total_pages (int): Tổng số pages
            
        Returns:
            tuple: (success: bool, error_msg: str)
        """
        try:
            logger.info(f"Đang export page {page_num}/{total_pages} bằng export button")
            
            # Select all trên page hiện tại
            headers_list = self.wait.until(
                EC.presence_of_all_elements_located(config.DOCUMENT_HEADERS_LOCATOR)
            )
            select_all_checkbox = headers_list[-1]
            select_all_checkbox.click()
            time.sleep(1)
            
            # Click download button
            download_document_btns = self.wait.until(
                EC.presence_of_all_elements_located(config.DOCUMENT_ACTION_BTNS_LOCALTOR)
            )
            download_document_btn = download_document_btns[0]
            download_document_btn.click()
            time.sleep(1)
            
            # Thử click export button
            try:
                export_document_btns = self.wait.until(
                    EC.presence_of_all_elements_located(config.EXPORT_DOCUMENT_BTNS_LOCALTOR)
                )
                export_document_btn = export_document_btns[0]
                
                # Kiểm tra button có bị disable không
                btn_class = export_document_btn.get_attribute('class') or ''
                if 'disabled' in btn_class:
                    raise Exception("Export button bị disable")
                
                export_document_btn.click()
                time.sleep(1)
                
                # Click OK
                btn_ok = self.wait.until(
                    EC.element_to_be_clickable(config.OK_BTN_LOCALTOR)
                )
                btn_ok.click()
                time.sleep(15)
                select_all_checkbox.click()
                time.sleep(2)
                
                # Chuyển sang page tiếp theo (nếu chưa phải page cuối)
                if page_num < total_pages:
                    try:
                        next_page_btn = self.wait.until(
                            EC.element_to_be_clickable(config.NEXT_PAGE_BTN_LOCATOR)
                        )
                        if 'disabled' not in next_page_btn.get_attribute('class'):
                            next_page_btn.click()
                            logger.info(f"Đã chuyển sang page {page_num + 1}")
                            time.sleep(4)
                        else:
                            logger.warning("Nút next page bị disable")
                            return False, "Nút next page bị disable"
                    except TimeoutException:
                        logger.warning("Không tìm thấy nút next page")
                        return False, "Không tìm thấy nút next page"
                
                # Chờ zip file được tải về
                success, zip_file_path = self._wait_for_file_download(expected_extension=".zip", timeout=300)
                if not success:
                    error_msg = f"Không tìm thấy zip file sau khi download page {page_num}"
                    logger.error(error_msg)
                    return False, error_msg
                
                # Kiểm tra số lượng file
                file_count = self._count_files_in_download_dir()
                if file_count > 1:
                    error_msg = f"Phát hiện {file_count} file trong download_dir, chỉ mong đợi 1 file zip"
                    logger.warning(error_msg)
                    return False, error_msg
                
                # Tạo folder zip cho client nếu chưa có
                zip_client_folder_name = "_".join([
                    str(client_object.client_name), 
                    str(client_object.client_number), 
                    "zip"
                ])
                zip_client_folder_path = os.path.join(self.zip_dir, zip_client_folder_name)
                os.makedirs(zip_client_folder_path, exist_ok=True)
                
                # Move zip vào storage
                success, zip_dest_path, error_msg = move_zip_to_storage(
                    zip_file_path, zip_client_folder_path
                )
                if not success:
                    logger.error(error_msg)
                    return False, error_msg
                
                # Giải nén zip
                success, error_msg = extract_zip(zip_dest_path, zip_client_folder_path)
                if not success:
                    logger.error(error_msg)
                    return False, error_msg
                
                # Xóa file zip sau khi giải nén
                remove_file(zip_dest_path)
                
                logger.info(f"Đã export page {page_num} thành công bằng export button")
                return True, ""
                
            except (TimeoutException, NoSuchElementException, Exception) as e:
                # Không thể click export button, chuyển sang export từng file
                download_document_btn.click()
                select_all_checkbox.click()
                error_msg = f"Không thể click export button cho page {page_num}: {str(e)}"
                logger.warning(error_msg)
                logger.info(f"Chuyển sang export từng file riêng lẻ cho page {page_num}")
                
                # Export từng file lẻ cho page này
                # Lấy danh sách documents trên page hiện tại
                document_table = self.wait.until(
                    EC.presence_of_element_located(config.DOCUMENT_TABLE_LOCATOR)
                )
                document_rows = document_table.find_elements(
                    config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                    config.DOCUMENT_TABLE_DIV_LOCATOR[1]
                )
                
                if len(document_rows) <= 0:
                    return False, "Không tìm thấy document row trong table"
                
                # Export từng file
                for i, row in enumerate(document_rows):
                    try:
                        document_data_cells = row.find_elements(
                            config.DOCUMENT_DATA_CELL_LOCATOR[0], 
                            config.DOCUMENT_DATA_CELL_LOCATOR[1]
                        )
                        
                        if len(document_data_cells) < 10:
                            continue
                        
                        row_doc_id = document_data_cells[9].text.strip()
                        
                        # Tìm document object
                        document = None
                        for doc in client_object.document_list:
                            if doc.document_id == row_doc_id:
                                document = doc
                                break
                        
                        if not document:
                            logger.warning(f"Không tìm thấy document object cho doc_id: {row_doc_id}")
                            continue
                        
                        # Export file
                        success, file_path, error_msg = self.click_export_single_file(i, document)
                        if success:
                            # Cập nhật Excel
                            download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            row_index = None  # TODO: Lấy từ doc_id_to_row_index nếu có
                            if self.excel_handler and row_index:
                                self.excel_handler.log_document_success(
                                    document.document_id, row_index, 
                                    document.document_name_with_id, 
                                    file_path, document.category_name, download_time
                                )
                            document.set_download_status("Success", "File downloaded successfully", download_time)
                            self.downloaded_documents += 1
                        else:
                            logger.warning(f"Không thể export file cho document {row_doc_id}: {error_msg}")
                            if self.excel_handler and row_index:
                                self.excel_handler.log_document_error(document.document_id, row_index, error_msg)
                            document.set_download_status("Error", error_msg)
                    
                    except Exception as e:
                        logger.error(f"Lỗi khi export file cho row {i}: {str(e)}")
                        continue
                
                # Chuyển sang page tiếp theo (nếu chưa phải page cuối)
                if page_num < total_pages:
                    try:
                        next_page_btn = self.wait.until(
                            EC.element_to_be_clickable(config.NEXT_PAGE_BTN_LOCATOR)
                        )
                        if 'disabled' not in next_page_btn.get_attribute('class'):
                            next_page_btn.click()
                            logger.info(f"Đã chuyển sang page {page_num + 1}")
                            time.sleep(4)
                        else:
                            logger.warning("Nút next page bị disable")
                            return False, "Nút next page bị disable"
                    except TimeoutException:
                        logger.warning("Không tìm thấy nút next page")
                        return False, "Không tìm thấy nút next page"
                
                return True, ""
                
        except Exception as e:
            error_msg = f"Lỗi khi export page {page_num}: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            return False, error_msg
    
    def export_page_individual_files(self, client_name, csv_documents_dict, 
                                     doc_id_to_row_index, client_target_dir):
        """
        Export từng file riêng lẻ cho page hiện tại (khi export button bị disable)
        Method này được gọi từ export_multiple khi không thể click export button
        """
        try:
            logger.info(f"Đang export từng file riêng lẻ cho page hiện tại của client: {client_name}")
            
            # Tìm document row trong table
            document_table = self.wait.until(
                EC.presence_of_element_located(config.DOCUMENT_TABLE_LOCATOR)
            )
            document_rows = document_table.find_elements(
                config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                config.DOCUMENT_TABLE_DIV_LOCATOR[1]
            )

            if len(document_rows) <= 0:
                error_msg = "Không tìm thấy document row trong table"
                logger.error(error_msg)
                self.error_description = error_msg
                return False
            
            # Lặp qua từng row
            for i, row in enumerate(document_rows):
                try:
                    logger.info(f"Đang xử lý document row {i+1}/{len(document_rows)}")
                    
                    document_data_cells = row.find_elements(
                        config.DOCUMENT_DATA_CELL_LOCATOR[0], 
                        config.DOCUMENT_DATA_CELL_LOCATOR[1]
                    )

                    if len(document_data_cells) < 10:
                        error_msg = f"Không tìm thấy đủ cells trong document row {i+1}"
                        logger.warning(error_msg)
                        continue
                    
                    row_doc_id = document_data_cells[9].text.strip()
                    if row_doc_id not in csv_documents_dict.keys():
                        error_msg = f"Document ID '{row_doc_id}' không hợp lệ (không có trong CSV)"
                        logger.warning(error_msg)
                        continue
                    
                    # Lấy thông tin document từ csv_documents_dict
                    doc_info = csv_documents_dict[row_doc_id]
                    expected_file_name = doc_info.get("expected_download_file_name", "")
                    doc_year = doc_info.get("Year", "").strip()
                    
                    # Xác định category folder
                    file_section = doc_info.get("File Section", "").strip()
                    document_type = doc_info.get("Document Type", "").strip()
                    description = doc_info.get("Description", "").strip()
                    category_folder = get_document_category(file_section, document_type, description)
                    category_dir = os.path.join(client_target_dir, category_folder)
                    
                    # Kiểm tra file đã tồn tại chưa (kiểm tra cả tên gốc và tên có doc_id)
                    if doc_year:
                        year_dir_status, year_dir = self._create_folder(doc_year, category_dir)
                        if not year_dir_status:
                            error_msg = f"Không thể tạo year folder: {doc_year}"
                            logger.error(error_msg)
                            row_index = doc_id_to_row_index.get(row_doc_id)
                            if row_index:
                                self.update_document_status_in_excel(
                                    row_doc_id, row_index, "Error", error_msg
                                )
                            continue
                        expected_base, expected_ext = os.path.splitext(expected_file_name)
                        file_path_with_doc_id = os.path.join(category_dir, doc_year, f"{expected_base}_{row_doc_id}{expected_ext}")
                    else:
                        expected_base, expected_ext = os.path.splitext(expected_file_name)
                        file_path_with_doc_id = os.path.join(category_dir, f"{expected_base}_{row_doc_id}{expected_ext}")
                    
                    # Kiểm tra file đã tồn tại (ưu tiên tên có doc_id)
                    existing_file_path = None
                    if os.path.exists(file_path_with_doc_id):
                        existing_file_path = file_path_with_doc_id
                    
                    if existing_file_path:
                        logger.info(f"File đã tồn tại: {existing_file_path}")
                        self.downloaded_documents += 1
                        download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            existing_file_name = os.path.basename(existing_file_path)
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Success", "File already exists",
                                existing_file_name, existing_file_path, download_time
                            )
                        continue
                    
                    # Click export button cho row này
                    document_first_cell = row.find_elements(
                        config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[0], 
                        config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[1]
                    )
                    if not document_first_cell:
                        error_msg = f"Không tìm thấy first cell của document row {i+1}"
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                    btns = document_first_cell[0].find_elements(By.TAG_NAME, "button")
                    if len(btns) < 3:
                        error_msg = f"Không tìm thấy export button cho document row {i+1}"
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                    btn_export = btns[2]
                    
                    # Kiểm tra button có bị disable không
                    if not btn_export.is_enabled():
                        error_msg = f"Export button bị disable cho document {row_doc_id}"
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                    btn_export.click()
                    logger.info(f"Đã click export button cho document {row_doc_id}")
                    time.sleep(5)
                    
                    # Chờ file được tải về
                    success, downloaded_file_path = self._wait_for_file_download(timeout=120)
                    if not success:
                        error_msg = f"Không tìm thấy file sau khi download cho document {row_doc_id}"
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                    # Kiểm tra số lượng file
                    file_count = self._count_files_in_download_dir()
                    if file_count > 1:
                        error_msg = f"Phát hiện {file_count} file trong download_dir, chỉ mong đợi 1 file cho document {row_doc_id}"
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Warning", error_msg
                            )
                        # Xóa file để tiếp tục
                        try:
                            os.remove(downloaded_file_path)
                        except:
                            pass
                        continue
                    
                    # Kiểm tra tên file
                    downloaded_file_name = os.path.basename(downloaded_file_path)
                    base_name, ext = os.path.splitext(downloaded_file_name)
                    expected_base, expected_ext = os.path.splitext(expected_file_name)
                    
                    if base_name != expected_base or ext != expected_ext:
                        error_msg = (f"Tên file không khớp cho document {row_doc_id}: "
                                   f"expected '{expected_file_name}', got '{downloaded_file_name}'")
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Warning", error_msg
                            )
                        # Vẫn tiếp tục move file
                    
                    # Đổi tên file để thêm document_id
                    new_file_name = f"{base_name}_{row_doc_id}{ext}"
                    renamed_file_path = os.path.join(category_dir, new_file_name)
                    try:
                        os.rename(downloaded_file_path, renamed_file_path)
                    except Exception as e:
                        error_msg = f"Lỗi khi đổi tên file cho document {row_doc_id}: {str(e)}"
                        logger.error(error_msg)
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                    # Xác định destination path
                    if doc_year:
                        final_dest = os.path.join(category_dir, doc_year, new_file_name)
                    else:
                        final_dest = os.path.join(category_dir, new_file_name)
                    
                    # Move file vào category hoặc category/year folder
                    try:
                        shutil.move(renamed_file_path, final_dest)
                        logger.info(f"Đã move file vào: {final_dest}")
                        self.downloaded_documents += 1
                        
                        # Cập nhật Excel
                        download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Success", "File downloaded successfully",
                                new_file_name, final_dest, download_time
                            )
                        
                    except Exception as e:
                        error_msg = f"Lỗi khi move file cho document {row_doc_id}: {str(e)}"
                        logger.error(error_msg)
                        logger.error(traceback.format_exc())
                        row_index = doc_id_to_row_index.get(row_doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                row_doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                except Exception as e:
                    error_msg = f"Lỗi khi xử lý document row {i+1}: {str(e)}"
                    logger.error(error_msg)
                    logger.error(traceback.format_exc())
                    # Cố gắng lấy doc_id để log
                    try:
                        document_data_cells = row.find_elements(
                            config.DOCUMENT_DATA_CELL_LOCATOR[0], 
                            config.DOCUMENT_DATA_CELL_LOCATOR[1]
                        )
                        if len(document_data_cells) >= 10:
                            row_doc_id = document_data_cells[9].text.strip()
                            if row_doc_id in doc_id_to_row_index:
                                row_index = doc_id_to_row_index.get(row_doc_id)
                                if row_index:
                                    self.update_document_status_in_excel(
                                        row_doc_id, row_index, "Error", error_msg
                                    )
                    except:
                        pass
                    continue
            
            logger.info(f"Hoàn thành export từng file riêng lẻ cho page hiện tại")
            return True
            
        except Exception as e:
            error_msg = f"Lỗi tổng quát khi export từng file riêng lẻ: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False
    
    def export_multiple(self, client_name, client_number, total_documents, 
                       csv_documents_dict, doc_id_to_row_index, client_target_dir):
        """Export multiple documents theo từng page (khi total_documents > 1)"""
        try:
            logger.info(f"Đang export multiple documents cho client: {client_name} (Tổng: {total_documents})")
            
            number_items_per_page = int(self.config.get('NUMBER_ITEMS_PER_PAGE', 50))
            total_pages = (total_documents + number_items_per_page - 1) // number_items_per_page
            
            # Tạo folder zip cho client
            zip_client_folder_name = "_".join([str(client_name), str(client_number), "zip"])
            zip_client_folder_path = os.path.join(self.zip_dir, zip_client_folder_name)
            os.makedirs(zip_client_folder_path, exist_ok=True)
            
            # Xử lý từng page
            for page_num in range(1, total_pages + 1):
                logger.info(f"--- Đang xử lý page {page_num}/{total_pages} ---")
                
                try:
                    # Select all trên page hiện tại
                    headers_list = self.wait.until(
                        EC.presence_of_all_elements_located(config.DOCUMENT_HEADERS_LOCATOR)
                    )
                    select_all_checkbox = headers_list[-1]
                    select_all_checkbox.click()
                    time.sleep(1)
                    
                    # Click download button
                    download_document_btns = self.wait.until(
                        EC.presence_of_all_elements_located(config.DOCUMENT_ACTION_BTNS_LOCALTOR)
                    )
                    download_document_btn = download_document_btns[0]
                    download_document_btn.click()
                    time.sleep(1)
                    
                    # Thử click export button, nếu không được thì chuyển sang export từng file
                    try:
                        # Click export button
                        export_document_btns = self.wait.until(
                            EC.presence_of_all_elements_located(config.EXPORT_DOCUMENT_BTNS_LOCALTOR)
                        )
                        export_document_btn = export_document_btns[0]
                        
                        # Kiểm tra button có bị disable không (kiểm tra class có chứa "disabled")
                        btn_class = export_document_btn.get_attribute('class') or ''
                        if 'disabled' in btn_class:
                            raise Exception("Export button bị disable")
                        
                        export_document_btn.click()
                        time.sleep(1)
                        
                    except (TimeoutException, NoSuchElementException, Exception) as e:
                        download_document_btn.click()
                        select_all_checkbox.click()
                        error_msg = f"Không thể click export button cho page {page_num}: {str(e)}"
                        logger.warning(error_msg)
                        logger.info(f"Chuyển sang export từng file riêng lẻ cho page {page_num}")
                        individual_export_success = self.export_page_individual_files(
                            client_name, csv_documents_dict, 
                            doc_id_to_row_index, client_target_dir
                        )
                        if not individual_export_success:
                            logger.warning(f"Export từng file riêng lẻ cho page {page_num} có một số lỗi")
                        
                        # Chuyển sang page tiếp theo (nếu chưa phải page cuối)
                        if page_num < total_pages:
                            try:
                                next_page_btn = self.wait.until(
                                    EC.element_to_be_clickable(config.NEXT_PAGE_BTN_LOCATOR)
                                )
                                if 'disabled' not in next_page_btn.get_attribute('class'):
                                    next_page_btn.click()
                                    logger.info(f"Đã chuyển sang page {page_num + 1}")
                                    time.sleep(4)
                                else:
                                    logger.warning("Nút next page bị disable")
                                    break
                            except TimeoutException:
                                logger.warning("Không tìm thấy nút next page")
                                break
                        # Tiếp tục với page tiếp theo
                        continue
                    
                    # Click OK
                    btn_ok = self.wait.until(
                        EC.element_to_be_clickable(config.OK_BTN_LOCALTOR)
                    )
                    btn_ok.click()
                    time.sleep(15)
                    select_all_checkbox.click()
                    time.sleep(2)
                    # Chuyển sang page tiếp theo (nếu chưa phải page cuối)
                    if page_num < total_pages:
                        try:
                            next_page_btn = self.wait.until(
                                EC.element_to_be_clickable(config.NEXT_PAGE_BTN_LOCATOR)
                            )
                            if 'disabled' not in next_page_btn.get_attribute('class'):
                                next_page_btn.click()
                                logger.info(f"Đã chuyển sang page {page_num + 1}")
                                time.sleep(4)
                            else:
                                logger.warning("Nút next page bị disable")
                                break
                        except TimeoutException:
                            logger.warning("Không tìm thấy nút next page")
                            break
                    
                    # Chờ zip file được tải về
                    success, zip_file_path = self._wait_for_file_download(expected_extension=".zip", timeout=300)
                    if not success:
                        error_msg = f"Không tìm thấy zip file sau khi download page {page_num}"
                        logger.error(error_msg)
                        self.error_description = error_msg
                        continue
                    
                    # Kiểm tra số lượng file
                    file_count = self._count_files_in_download_dir()
                    if file_count > 1:
                        error_msg = f"Phát hiện {file_count} file trong download_dir, chỉ mong đợi 1 file zip"
                        logger.warning(error_msg)
                        self.error_description = error_msg
                        continue
                    
                    # Move zip vào 0_zip_/client_name_client_number_zip/
                    zip_filename = os.path.basename(zip_file_path)
                    zip_dest_path = os.path.join(zip_client_folder_path, zip_filename)
                    shutil.move(zip_file_path, zip_dest_path)
                    logger.info(f"Đã move zip vào: {zip_dest_path}")
                    
                    # Giải nén zip
                    try:
                        with zipfile.ZipFile(zip_dest_path, 'r') as zip_ref:
                            # Giải nén vào folder zip_client_folder_path
                            zip_ref.extractall(zip_client_folder_path)
                        logger.info(f"Đã giải nén zip: {zip_dest_path}")
                        
                        # Xóa file zip sau khi giải nén
                        os.remove(zip_dest_path)
                        
                    except zipfile.BadZipFile:
                        error_msg = f"Zip file bị hỏng: {zip_dest_path}"
                        logger.error(error_msg)
                        self.error_description = error_msg
                        continue
                    except Exception as e:
                        error_msg = f"Lỗi khi giải nén zip: {str(e)}"
                        logger.error(error_msg)
                        logger.error(traceback.format_exc())
                        self.error_description = error_msg
                        continue
                    
                except Exception as e:
                    error_msg = f"Lỗi khi xử lý page {page_num}: {str(e)}"
                    logger.error(error_msg)
                    logger.error(traceback.format_exc())
                    self.error_description = error_msg
                    continue
            
            # Sau khi tải xong tất cả các page, move từng file từ zip folder vào client folder
            logger.info("Đang move files từ zip folder vào client folder...")
            
            for doc_id, doc_info in csv_documents_dict.items():
                try:
                    expected_file_name = doc_info.get("expected_download_file_name", "")
                    doc_year = doc_info.get("Year", "").strip()
                    
                    # Tìm file trong zip folder (file có tên chứa document_id ở cuối)
                    # Format: expected_name_document_id.ext
                    found_file = None
                    for root, dirs, files in os.walk(zip_client_folder_path):
                        for file in files:
                            # File từ zip có format: expected_name_document_id.ext
                            if doc_id in file and expected_file_name.replace(".", f"_{doc_id}.") in file:
                                found_file = os.path.join(root, file)
                                break
                        if found_file:
                            break
                    
                    if not found_file:
                        # Không tìm thấy file
                        error_msg = f"Không tìm thấy file cho document_id {doc_id}"
                        logger.warning(error_msg)
                        row_index = doc_id_to_row_index.get(doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                doc_id, row_index, "Error", error_msg
                            )
                        continue
                    
                    # Xác định category folder dựa trên File Section, Document Type, Description
                    file_section = doc_info.get("File Section", "").strip()
                    document_type = doc_info.get("Document Type", "").strip()
                    description = doc_info.get("Description", "").strip()
                    category_folder = get_document_category(file_section, document_type, description)
                    logger.info(f"Document {doc_id} category: {category_folder} (File Section: {file_section}, Document Type: {document_type})")
                    
                    # get category folder path
                    category_dir = os.path.join(client_target_dir, category_folder)
                    
                    # get year folder path
                    if doc_year:
                        year_dir_status, year_dir = self._create_folder(doc_year, category_dir)
                        if not year_dir_status:
                            error_msg = f"Không thể tạo year folder: {doc_year}"
                            logger.error(error_msg)
                            row_index = doc_id_to_row_index.get(doc_id)
                            if row_index:
                                self.update_document_status_in_excel(
                                    doc_id, row_index, "Error", error_msg
                                )
                            continue
                        final_dest = os.path.join(year_dir, expected_file_name)
                    else:
                        final_dest = os.path.join(category_dir, expected_file_name)
                    
                    # Kiểm tra file đã tồn tại chưa
                    if os.path.exists(final_dest):
                        logger.info(f"File đã tồn tại: {final_dest}")
                        self.downloaded_documents += 1
                        download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        row_index = doc_id_to_row_index.get(doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                doc_id, row_index, "Success", "File already exists",
                                expected_file_name, final_dest, download_time
                            )
                        # Xóa file trong zip folder
                        try:
                            os.remove(found_file)
                        except:
                            pass
                        continue
                    
                    # Move file
                    try:
                        shutil.move(found_file, final_dest)
                        logger.info(f"Đã move file: {final_dest}")
                        self.downloaded_documents += 1
                        
                        # Cập nhật Excel
                        download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        row_index = doc_id_to_row_index.get(doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                doc_id, row_index, "Success", "File downloaded successfully",
                                expected_file_name, final_dest, download_time
                            )
                    except Exception as e:
                        error_msg = f"Lỗi khi move file: {str(e)}"
                        logger.error(error_msg)
                        logger.error(traceback.format_exc())
                        row_index = doc_id_to_row_index.get(doc_id)
                        if row_index:
                            self.update_document_status_in_excel(
                                doc_id, row_index, "Error", error_msg
                            )
                        continue
                        
                except Exception as e:
                    error_msg = f"Lỗi khi xử lý document {doc_id}: {str(e)}"
                    logger.error(error_msg)
                    logger.error(traceback.format_exc())
                    row_index = doc_id_to_row_index.get(doc_id)
                    if row_index:
                        self.update_document_status_in_excel(
                            doc_id, row_index, "Error", error_msg
                        )
                    continue
            
            logger.info(f"Hoàn thành export multiple: {self.downloaded_documents}/{total_documents} files")
            return True
            
        except Exception as e:
            error_msg = f"Lỗi tổng quát khi export multiple: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            self.error_description = error_msg
            return False
    
    def _reload_page(self):
        """Reload trang hiện tại"""
        try:
            logger.info("Đang reload trang...")
            self.driver.refresh()
            time.sleep(5)
            logger.info("Đã reload trang")
            return True
        except Exception as e:
            logger.error(f"Lỗi khi reload trang: {str(e)}")
            return False
    
    def _is_web_error(self, error_msg):
        """
        Kiểm tra xem lỗi có phải là lỗi web không (cần reload và retry)
        
        Args:
            error_msg (str): Error message
            
        Returns:
            bool: True nếu là lỗi web
        """
        web_error_keywords = [
            "timeout", "timed out", "element not found", "no such element",
            "stale element", "element is not attached", "session not created",
            "connection", "network", "webdriver", "selenium"
        ]
        error_lower = error_msg.lower()
        return any(keyword in error_lower for keyword in web_error_keywords)
    
    def process_client(self, client_info, max_retries=2):
        """
        Xử lý một client với retry logic
        
        Args:
            client_info (dict): Thông tin client
            max_retries (int): Số lần retry tối đa khi gặp lỗi web
            
        Returns:
            bool: True nếu thành công, False nếu có lỗi
        """
        retry_count = 0
        while retry_count <= max_retries:
            try:
                client_name = client_info['client_name']
                client_number = client_info['client_number']
                logger.info(f"Bắt đầu xử lý client: {client_name} ({client_number})")
                
                # Reset tracking variables
                self.total_documents = 0
                self.downloaded_documents = 0
                self.error_description = ""
                
                # Dọn sạch download_dir
                if not self._clean_download_dir():
                    error_msg = "Không thể dọn sạch download_dir"
                    logger.error(error_msg)
                    self.error_description = error_msg
                    return False
                
                # Tìm kiếm client
                if not self.search_client(client_name):
                    error_msg = f"Không thể tìm kiếm client: {self.error_description}"
                    logger.error(error_msg)
                    return False
                
                # Kiểm tra client tồn tại
                client_exists, total_documents = self.check_client_exists(client_name, client_number)
                if not client_exists:
                    error_msg = f"Client không tồn tại: {self.error_description}"
                    logger.error(error_msg)
                    return False
                
                self.total_documents = total_documents or 0
                
                # Xử lý theo số lượng documents
                if self.total_documents == 0:
                    error_msg = "Client has no document"
                    logger.warning(error_msg)
                    self.error_description = error_msg
                    return False
                
                # Tải CSV list
                csv_success, csv_file_path = self.download_csv_list(client_name, client_number)
                if not csv_success:
                    error_msg = f"Không thể tải CSV list: {self.error_description}"
                    logger.error(error_msg)
                    return False
                
                # Đọc CSV file
                csv_documents_dict = self.read_csv_file(csv_file_path, client_name, client_number)
                if not csv_documents_dict:
                    error_msg = f"Không thể đọc CSV file: {self.error_description}"
                    logger.error(error_msg)
                    return False
                
                # Log documents vào Excel
                doc_id_to_row_index = self.log_documents_to_excel(
                    csv_documents_dict, client_name, client_number
                )
                if not doc_id_to_row_index:
                    error_msg = "Không thể log documents vào Excel"
                    logger.error(error_msg)
                    self.error_description = error_msg
                    return False
                
                client_dir = "-".join([str(client_name), str(client_number)])
                client_dir_status, client_target_dir = self._get_safe_client_dir(client_dir, self.download_dir, is_client_folder=True)
                if not client_dir_status:
                    error_msg = "Không thể tạo client folder"
                    logger.error(error_msg)
                    self.error_description = error_msg
                    return False
                
                # Export documents
                if self.total_documents == 1:
                    export_success = self.export_single(
                        client_name, client_number, csv_documents_dict, doc_id_to_row_index, client_target_dir
                    )
                else:
                    export_success = self.export_multiple(
                        client_name, client_number, self.total_documents,
                        csv_documents_dict, doc_id_to_row_index, client_target_dir
                    )
                
                if not export_success:
                    error_msg = f"Không thể export documents: {self.error_description}"
                    logger.error(error_msg)
                    return False
                
                logger.info(f"Hoàn thành xử lý client: {client_name} ({self.downloaded_documents}/{self.total_documents} files)")
                return True
                
            except Exception as e:
                error_msg = f"Lỗi tổng quát khi xử lý client: {str(e)}"
                logger.error(error_msg)
                logger.error(traceback.format_exc())
                self.error_description = error_msg
                
                # Kiểm tra xem có phải lỗi web không
                if self._is_web_error(error_msg) and retry_count < max_retries:
                    retry_count += 1
                    logger.warning(f"Gặp lỗi web, đang retry lần {retry_count}/{max_retries}...")
                    
                    # Reload trang
                    if self._reload_page():
                        # Thử lại từ đầu
                        continue
                    else:
                        logger.error("Không thể reload trang, bỏ qua retry")
                        return False
                else:
                    # Không phải lỗi web hoặc đã hết retry
                    return False
    
    def run(self, excel_file_path):
        """Chạy automation cho tất cả client"""
        try:
            logger.info("=" * 80)
            logger.info("BẮT ĐẦU AUTOMATION")
            logger.info("=" * 80)
            
            # Setup driver (chỉ 1 lần cho tất cả client)
            self.setup_driver()
            
            # Đọc danh sách client
            client_list = self.read_client_list(excel_file_path)
            if not client_list:
                logger.warning("Không có client nào có Status = Pending")
                return
            
            # Mở trang GOFILEROOM
            self.driver.get(config.BASE_URL)
            time.sleep(10)
            
            # Kiểm tra URL sau khi load để xem có bị chuyển hướng sang trang login không
            current_url = self.driver.current_url
            logger.info(f"URL hiện tại sau khi mở trang: {current_url}")
            
            if config.LOGIN_URL in current_url:
                logger.warning(f"Trang web bị chuyển hướng sang trang login: {current_url}")
                logger.info("Đang thực hiện login tự động...")
                
                # Thực hiện login
                login_success = self.login()
                if not login_success:
                    error_msg = f"Không thể login vào GOFILEROOM: {self.error_description}"
                    logger.error(error_msg)
                    raise Exception(error_msg)
                
                # Kiểm tra lại URL sau khi login
                current_url = self.driver.current_url
                logger.info(f"URL sau khi login: {current_url}")
            
            if config.BASE_URL not in current_url and config.LOGIN_URL not in current_url:
                logger.warning(f"URL hiện tại ({current_url}) khác với BASE_URL ({config.BASE_URL})")
                logger.warning("Tiếp tục xử lý nhưng có thể có vấn đề...")
            
            # Xử lý từng client
            success_count = 0
            total_count = len(client_list)
            consecutive_errors = 0
            max_consecutive_errors = 10
            error_summary = {}
            
            for idx, client_info in enumerate(client_list, 1):
                logger.info(f"\n{'=' * 80}")
                logger.info(f"Xử lý client {idx}/{total_count}: {client_info['client_name']}")
                logger.info(f"{'=' * 80}\n")
                
                try:
                    # Process client
                    process_success = self.process_client(client_info)
                    
                    # Reset consecutive_errors nếu thành công
                    if process_success:
                        consecutive_errors = 0
                        error_summary = {}
                    
                    # Cập nhật Excel
                    if process_success:
                        if self.total_documents == self.downloaded_documents:
                            client_info['status_cell'].value = "Success"
                            client_info['description_cell'].value = "Download client documents successfully"
                        else:
                            client_info['status_cell'].value = "Warning"
                            client_info['description_cell'].value = (
                                f"Downloaded {self.downloaded_documents}/{self.total_documents} files. "
                                f"{self.error_description}"
                            )
                        success_count += 1
                    else:
                        # Tăng consecutive_errors nếu là lỗi nghiêm trọng (không phải warning)
                        if self.total_documents > 0:  # Có documents nhưng không download được
                            consecutive_errors += 1
                            error_summary[f"Client_{idx}"] = {
                                "client_name": client_info['client_name'],
                                "client_number": client_info['client_number'],
                                "error": self.error_description
                            }
                            
                            # Kiểm tra nếu đạt max consecutive errors
                            if consecutive_errors >= max_consecutive_errors:
                                logger.error(f"Đã gặp {consecutive_errors} lỗi liên tiếp, dừng automation và gửi email")
                                
                                # Gửi email
                                if self.email_handler:
                                    self.email_handler.send_critical_error_email(
                                        consecutive_errors, error_summary
                                    )
                                
                                # Cập nhật Excel cho client hiện tại
                                client_info['status_cell'].value = "Error"
                                client_info['description_cell'].value = (
                                    f"{self.error_description} "
                                    f"(Automation stopped due to {consecutive_errors} consecutive errors)"
                                )
                                self.excel_handler.save_workbook()
                                
                                raise Exception(
                                    f"Automation stopped: {consecutive_errors} consecutive errors. "
                                    f"Email notification sent."
                                )
                        
                        if self.total_documents == 0:
                            client_info['status_cell'].value = "Warning"
                            client_info['description_cell'].value = self.error_description
                        else:
                            client_info['status_cell'].value = "Error"
                            client_info['description_cell'].value = self.error_description
                    
                    # Cập nhật các cột khác
                    client_info['total_documents_cell'].value = str(self.total_documents)
                    client_info['num_files_downloaded_cell'].value = str(self.downloaded_documents)
                    
                    # Cập nhật client folder path
                    client_dir = "-".join([str(client_info['client_name']), str(client_info['client_number'])])
                    _, client_folder_path = self._get_safe_client_dir(client_dir, self.download_dir, is_client_folder=True)
                    client_info['client_folder_path_cell'].value = client_folder_path
                    
                    # Lưu Excel sau mỗi client
                    self.excel_handler.save_workbook()
                    logger.info(f"Đã lưu Excel sau khi xử lý client {idx}")
                    
                    # Nghỉ giữa các client
                    if idx < total_count:
                        time.sleep(3)
                    
                except Exception as e:
                    error_msg = f"Lỗi khi xử lý client {idx}: {str(e)}"
                    logger.error(error_msg)
                    logger.error(traceback.format_exc())
                    client_info['status_cell'].value = "Error"
                    client_info['description_cell'].value = error_msg
                    try:
                        self.workbook.save(excel_file_path)
                    except:
                        pass
                    continue
            
            logger.info(f"\n{'=' * 80}")
            logger.info(f"HOÀN THÀNH AUTOMATION")
            logger.info(f"Thành công: {success_count}/{total_count}")
            logger.info(f"{'=' * 80}\n")
            
        except Exception as e:
            logger.error(f"Lỗi tổng quát trong quá trình automation: {str(e)}")
            logger.error(traceback.format_exc())
            try:
                if self.workbook:
                    self.workbook.save(excel_file_path)
            except:
                pass
            raise
        finally:
            self.cleanup()
    
    def cleanup(self):
        """Dọn dẹp và đóng browser"""
        try:
            if self.driver:
                self.driver.quit()
                logger.info("Đã đóng browser")
        except Exception as e:
            logger.error(f"Lỗi khi cleanup: {str(e)}")

def main():
    """Entry point"""
    try:
        downloader = GofileRoomDownloader()
        
        # Kiểm tra file Excel có tồn tại không
        if not os.path.exists(downloader.excel_file_path):
            logger.error(f"File Excel không tồn tại: {downloader.excel_file_path}")
            return
        
        downloader.run(downloader.excel_file_path)
        
    except Exception as e:
        logger.error(f"Lỗi chính: {str(e)}")
        logger.error(traceback.format_exc())


if __name__ == "__main__":
    main()
