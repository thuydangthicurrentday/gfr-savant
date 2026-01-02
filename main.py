"""
Main script for GOFILEROOM Downloader
Refactored version with new structure
"""

import os
import csv
import re
import time
import logging
import traceback
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException

import config
from document_mapping import get_document_category
from file_handler import (
    rename_file_with_doc_id, move_file, move_csv_to_storage,
    move_zip_to_storage, extract_zip, remove_file, find_file_in_zip_folder,
    wait_for_file_download, clean_download_dir, rename_csv_file,
    FileHandlerError, FileNotFoundError, FileOperationError, ZipFileError, FileDownloadTimeoutError
)
from excel_handler import ExcelHandler, ExcelHandlerError, ExcelHeaderError, ExcelOperationError, ExcelSaveError
from models import Document, Client, BASE_DOWNLOAD_DIR
from utils import resource_path, load_env_config
from email_handler import create_email_handler_from_config

# Setup logging
log_file_name = 'gofileroom_download.log'
log_file_path = resource_path(log_file_name)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# Custom Exceptions
class GofileRoomDownloaderError(Exception):
    """Base exception cho GofileRoomDownloader errors"""
    pass


class ClientNotFoundError(GofileRoomDownloaderError):
    """Raised when client is not found"""
    pass


class CSVExportError(GofileRoomDownloaderError):
    """Raised when there is an error exporting CSV file"""
    pass


class DocumentExportError(GofileRoomDownloaderError):
    """Raised when there is an error exporting documents"""
    pass


class WebNavigationError(GofileRoomDownloaderError):
    """Raised when there is an error related to web navigation (redirect to login, etc.)"""
    pass


class LoginError(GofileRoomDownloaderError):
    """Raised when there is an error during login"""
    pass


class GofileRoomDownloader:
    def __init__(self):
        """Initialize downloader"""
        self.driver = None
        self.wait = None
        self.excel_handler = None
        
        # Load config from .env
        self.config = load_env_config(raise_on_not_found=True)
        
        # Initialize email handler
        self.email_handler = create_email_handler_from_config(self.config)
        
        # Download configuration
        self.download_dir = BASE_DOWNLOAD_DIR
        self.csv_dir = os.path.join(self.download_dir, "0_csv_")
        self.zip_dir = os.path.join(self.download_dir, "0_zip_")
        
        # Create necessary directories
        os.makedirs(self.download_dir, exist_ok=True)
        os.makedirs(self.csv_dir, exist_ok=True)
        os.makedirs(self.zip_dir, exist_ok=True)
        
        # Excel file properties
        self.excel_file_name = self.config.get('CLIENT_LIST_FILE_NAME', 'download_gofileroom_data.xlsx')
        self.excel_file_path = resource_path(self.excel_file_name)
        
        logger.info("GofileRoomDownloader initialized")
        
        # Setup driver and load home page
        self.setup_driver()
        self.load_home_page()

    def setup_driver(self):
        """Setup Chrome WebDriver - connect to running Chrome instance"""
        logger.info("Connecting to running Chrome...")
        
        chrome_options = Options()
        prefs = {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        logger.info("Successfully connected to running Chrome")
        
        self.driver.implicitly_wait(config.IMPLICIT_WAIT)
        self.driver.set_page_load_timeout(config.PAGE_LOAD_TIMEOUT)
        self.wait = WebDriverWait(self.driver, config.EXPLICIT_WAIT)

    def load_home_page(self):
        """Load home page, if redirected to login page then login again"""
        logger.info("Loading home page...")
        
        # Check current URL
        self.driver.get(config.BASE_URL)
        time.sleep(5)
        current_url = self.driver.current_url
        logger.info(f"Current URL: {current_url}")
        
        # If on login page, perform login
        if config.LOGIN_URL in current_url:
            logger.info("Detected login page, performing login...")
            self.login()
        
        # If not on home page, navigate to home page
        if config.BASE_URL not in current_url:
            self.driver.get(config.BASE_URL)
            time.sleep(3)
            
            # Check again if redirected to login
            current_url = self.driver.current_url
            if config.LOGIN_URL in current_url:
                logger.info("Redirected to login page after loading home page, performing login...")
                self.login()
        
        logger.info("Successfully loaded home page")

    def _check_and_handle_login_redirect(self):
        """
        Check if redirected to login page, if so then login
        
        Raises:
            WebNavigationError: If redirected to login page
        """
        current_url = self.driver.current_url
        if config.LOGIN_URL in current_url:
            error_msg = "Redirected to login page"
            logger.error(error_msg)
            raise WebNavigationError(error_msg)
    
    def login(self):
        """
        Perform login to GOFILEROOM
        
        Raises:
            LoginError: If there is an error during login
        """
        logger.info("Logging into GOFILEROOM...")
        
        # Read username from config
        username = self.config.get('USERNAME')
        
        if not username:
            error_msg = "USERNAME not found in .env file"
            logger.error(error_msg)
            raise LoginError(error_msg)
        
        # Switch to default context
        self.driver.switch_to.default_content()
        
        # Find and enter username
        try:
            try:
                username_input = self.wait.until(
                    EC.presence_of_element_located((By.ID, "txtLogin"))
                )
            except TimeoutException:
                username_input = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/form/div[1]/div/div/input"))
                )
        except TimeoutException as e:
            error_msg = "Timeout: Username input field not found"
            logger.error(error_msg)
            raise LoginError(error_msg) from e
        
        # Clear old content and enter username
        username_input.clear()
        username_input.send_keys(username)
        logger.info(f"Entered username: {username}")
        time.sleep(1)
        
        # Find and click Sign In button
        try:
            try:
                sign_in_btn = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "btnSignIn1"))
                )
            except TimeoutException:
                sign_in_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div[2]/form/button"))
                )
        except TimeoutException as e:
            error_msg = "Timeout: Sign In button not found"
            logger.error(error_msg)
            raise LoginError(error_msg) from e
        
        # Check if button is disabled
        btn_class = sign_in_btn.get_attribute('class') or ''
        if 'disabled' in btn_class or not sign_in_btn.is_enabled():
            error_msg = "Sign In button is disabled"
            logger.error(error_msg)
            raise LoginError(error_msg)
        
        sign_in_btn.click()
        logger.info("Clicked Sign In button")
        
        # Wait 15s for automatic redirect to BASE_URL
        logger.info("Waiting for redirect after login (15 seconds)...")
        time.sleep(15)
        
        # Check URL after login
        current_url = self.driver.current_url
        logger.info(f"URL after login: {current_url}")
        
        if config.LOGIN_URL in current_url:
            error_msg = f"Still on login page after login: {current_url}"
            logger.error(error_msg)
            raise LoginError(error_msg)
        
        logger.info("Login successful")

    def search_client(self, client_object):
        """
        Search for client and check if client exists on home page
        Get total documents from web page and assign to client_object.max_total_documents
        
        Args:
            client_object (Client): Client object
            
        Raises:
            ClientNotFoundError: If client is not found or there is an error during search
            WebNavigationError: If there is an error related to web navigation
        """
        try:
            logger.info(f"Searching for client: {client_object.client_name}")
            
            # Switch to default context
            self.driver.switch_to.default_content()
            
            # Find iframe containing search input
            iframe = self.wait.until(
                EC.presence_of_element_located(config.SEARCH_CLIENT_IFRAME_LOCATOR)
            )
            self.driver.switch_to.frame(iframe)
            
            # Switch to frame level 2
            FRAME_LEVEL_2 = (By.NAME, "mainFrame")
            self.wait.until(
                EC.frame_to_be_available_and_switch_to_it(FRAME_LEVEL_2)
            )

            search_input = None
            try:
                search_input = self.wait.until(
                    EC.element_to_be_clickable(config.SEARCH_INPUT_LOCATOR)
                )
            except TimeoutException:
                logger.warning("Not found with main selector, trying alternative selectors...")
                for selector in config.SEARCH_INPUT_ALTERNATIVES:
                    try:
                        search_input = self.driver.find_element(selector[0], selector[1])
                        if search_input.is_displayed() and search_input.is_enabled():
                            break
                        search_input = None
                    except NoSuchElementException:
                        continue
                
                if not search_input:
                    error_msg = "Search input field not found with any selector"
                    logger.error(error_msg)
                    raise WebNavigationError(error_msg)
            # Activate search input if needed
            if not search_input.is_enabled():
                try:
                    search_input.click()
                    time.sleep(1)
                except:
                    self.driver.execute_script("arguments[0].click();", search_input)
                    time.sleep(1)
            
            # Clear old content and enter client name
            search_input.clear()
            search_input.send_keys(client_object.client_name)
            search_input.send_keys(Keys.RETURN)
            
            logger.info(f"Entered client name '{client_object.client_name}' and pressed Enter")
            time.sleep(2)
            
            # Wait for client tree root element
            client_tree_root = self.wait.until(
                EC.presence_of_element_located(config.CLIENT_TREE_ROOT_LOCATOR)
            )
            time.sleep(1)
            
            client_list_ul = client_tree_root.find_element(By.XPATH, "./ul")
            client_item_a_tags = client_list_ul.find_elements(By.TAG_NAME, "a")
            
            if not client_item_a_tags:
                error_msg = "Client items not found in tree structure"
                logger.warning(error_msg)
                raise ClientNotFoundError(error_msg)
            
            # Find matching client
            client_item = None
            for a_tag in client_item_a_tags:
                a_tag_text = str(a_tag.text)
                search_text = f"{client_object.client_name} | {client_object.client_number}".lower()
                if a_tag_text.lower().startswith(search_text):
                    client_item = a_tag
                    break
            
            if not client_item:
                error_msg = f"Client '{client_object.client_name} | {client_object.client_number}' not found in tree structure"
                logger.warning(error_msg)
                raise ClientNotFoundError(error_msg)
            
            # Click on client to load document list
            client_item.click()
            logger.info("Clicked on client to load Document List")
            time.sleep(2)
            
            # Get number of documents from text (format: "Client Name | Number (count)")
            pattern = r'\((\d+)\)'
            match = re.search(pattern, client_item.text)
            if match:
                number_documents = int(match.group(1))
                logger.info(f"Found client with {number_documents} documents")
                client_object.max_total_documents = number_documents
            else:
                error_msg = "Document count not found in text"
                logger.warning(error_msg)
                raise ClientNotFoundError(error_msg)
                
        except (ClientNotFoundError, WebNavigationError):
            raise
        except Exception as e:
            error_msg = f"Error searching for client: {str(e)}"
            raise ClientNotFoundError(error_msg) from e

    def export_csv_file(self, client_object):
        """
        Download CSV file containing document list
        Download CSV file, move file to 0_csv_ directory
        Read document list from CSV file, create document objects and assign to client object
        For each document object, check if it already exists in client object's document list, if not then add it
        
        Args:
            client_object (Client): Client object
            
        Raises:
            CSVExportError: If there is an error exporting CSV file
            FileDownloadTimeoutError: If timeout waiting for CSV file download
            FileOperationError: If there is an error moving CSV file
        """
        try:
            logger.info("Downloading CSV file with document list...")
            
            # Click Export List button
            EXPORT_LIST_BTN_LOCATOR = (By.XPATH, "//button[contains(text(), 'Export List')]")
            btn_export_list = self.wait.until(
                EC.element_to_be_clickable(EXPORT_LIST_BTN_LOCATOR)
            )
            btn_export_list.click()
            logger.info("Clicked Export List button")
            
            # Wait for CSV file to download
            csv_file_path = wait_for_file_download(self.download_dir, expected_extension=".csv", timeout=120)
            
            # Move CSV file to 0_csv_ directory
            csv_dest_path = move_csv_to_storage(csv_file_path, self.csv_dir)
            logger.info(f"Moved CSV file to: {csv_dest_path}")
            
            # Read document list from CSV file (validate client match)
            documents_data = self._read_csv_file(csv_dest_path, client_object.client_name, client_object.client_number)
            
            # After validation, rename CSV file to: Search_<client_folder_name>.csv
            csv_dest_path = rename_csv_file(csv_dest_path, client_object.client_folder_name)
            logger.info(f"Renamed CSV file to: {os.path.basename(csv_dest_path)}")
            client_object.csv_download_file_path = csv_dest_path
            
            # Create document objects and add to client object
            for doc_id, doc_info in documents_data.items():
                document = Document(
                    document_id=doc_id,
                    file_section=doc_info.get("File Section", ""),
                    document_type=doc_info.get("Document Type", ""),
                    description=doc_info.get("Description", ""),
                    year=doc_info.get("Year", ""),
                    document_date=doc_info.get("Document Date", ""),
                    file_size=doc_info.get("File Size", ""),
                    file_type=doc_info.get("File Type", ""),
                    client_name=client_object.client_name,
                    client_object=client_object
                )
                
                # Check if document already exists in document_list, if not then add it
                client_object.add_document(document)
                
                # Add row to document log sheet (if not exists then add, if exists then skip)
                if self.excel_handler:
                    # Check if document already exists in Excel
                    existing_row_index = self.excel_handler.get_document_row_index(
                        document.client_name, 
                        document.client_object.client_number, 
                        document.document_id
                    )
                    
                    # If not exists, add new row
                    if not existing_row_index:
                        try:
                            row_index = self.excel_handler.add_document_row(
                                document.document_id,
                                document.client_name,
                                document.client_object.client_number,
                                file_name=document.document_name_with_id,
                                file_section=document.file_section,
                                document_type=document.document_type,
                                description=document.description,
                                year=document.year,
                                document_date=document.document_date,
                                file_size=document.file_size,
                                file_type=document.file_type,
                                folder_category=document.category_name
                            )
                            logger.debug(f"Added new document row to Excel: doc_id={document.document_id}, row={row_index}")
                        except (ExcelHeaderError, ExcelOperationError) as e:
                            logger.error(f"Error adding document row to Excel: {str(e)}")
                    else:
                        logger.debug(f"Document {document.document_id} already exists in Excel at row {existing_row_index}, skipping")
            
            logger.info(f"Created {len(client_object.document_list)} document objects for client {client_object.client_name}")
            
        except (CSVExportError, FileDownloadTimeoutError, FileOperationError):
            raise
        except Exception as e:
            error_msg = f"Error exporting CSV file: {str(e)}"
            raise CSVExportError(error_msg) from e

    def _read_csv_file(self, csv_file_path, client_name_check, client_number_check):
        """
        Read CSV file and return dictionary with Document ID as key
        
        Args:
            csv_file_path (str): CSV file path
            client_name_check (str): Client name for validation
            client_number_check (str): Client number for validation
            
        Returns:
            dict: Dictionary with Document ID as key, empty dict if only header exists
            
        Raises:
            CSVExportError: If CSV file does not exist, cannot be read, or does not match client
        """
        logger.info(f"Reading CSV file: {csv_file_path}")
        
        if not os.path.exists(csv_file_path):
            error_msg = f"CSV file does not exist: {csv_file_path}"
            logger.error(error_msg)
            raise CSVExportError(error_msg)
        
        documents_data = {}
        
        try:
            with open(csv_file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                
                # Read header
                try:
                    header = next(reader)
                    header = [h.strip() for h in header]
                    
                    # Find column indices
                    doc_id_index = header.index("Document ID")
                    client_name_index = header.index("Client Name")
                    client_number_index = header.index("Client Number")
                    
                except (StopIteration, ValueError) as e:
                    raise CSVExportError(f"Error reading CSV header: {str(e)}") from e
                
                # Read data
                try:
                    first_data_row = next(reader)
                except StopIteration:
                    logger.warning("CSV file only has header, no data")
                    return documents_data
                
                # Validate Client Name/Number from first row
                actual_client_name = first_data_row[client_name_index].strip().strip('"')
                actual_client_number = first_data_row[client_number_index].strip().strip('"')
                
                if actual_client_number != client_number_check:
                    error_msg = (f"CSV file does not match client: "
                               f"Expected '{client_name_check}'/'{client_number_check}', "
                               f"Got '{actual_client_name}'/'{actual_client_number}'")
                    logger.error(error_msg)
                    raise CSVExportError(error_msg)
                
                # Read first row
                doc_id = first_data_row[doc_id_index].strip().strip('"')
                documents_data[doc_id] = {
                    header[i]: first_data_row[i].strip().strip('"') for i in range(len(header))
                }
                
                # Read remaining rows
                for row in reader:
                    if not row:
                        continue
                    doc_id = row[doc_id_index].strip().strip('"')
                    documents_data[doc_id] = {
                        header[i]: row[i].strip().strip('"') for i in range(len(header))
                    }
            
            logger.info(f"Read {len(documents_data)} documents from CSV")
            return documents_data
            
        except CSVExportError:
            raise
        except Exception as e:
            error_msg = f"Error reading CSV file: {str(e)}"
            raise CSVExportError(error_msg) from e
    
    def export_documents(self, client_object):
        """
        Loop through all client document pages
        Each loop:
            - If redirected to login page then raise exception
            - If max_total_documents = 1 then perform download_single_document_in_page with row_index = 0
            - If max_total_documents > 1 then perform download_multiple_documents_in_page
                - If export button is disabled then perform download_single_document_in_page with row_index and break
            - Click next page button to load next page, if next page button not found then break
        
        Args:
            client_object (Client): Client object
            
        Raises:
            DocumentExportError: If there is an error exporting documents
            WebNavigationError: If redirected to login page
        """
        try:
            logger.info(f"Starting export documents for client: {client_object.client_name}")
            
            # Check if redirected to login page
            self._check_and_handle_login_redirect()
            
            # Calculate number of pages
            number_items_per_page = int(self.config.get('NUMBER_ITEMS_PER_PAGE', 50))
            total_pages = (client_object.max_total_documents + number_items_per_page - 1) // number_items_per_page
            
            # Get document table
            document_table = self.wait.until(
                EC.presence_of_element_located(config.DOCUMENT_TABLE_LOCATOR)
            )
            
            # Loop through each page
            for page_num in range(1, total_pages + 1):
                logger.info(f"--- Processing page {page_num}/{total_pages} ---")
                
                # Check again if redirected to login page
                self._check_and_handle_login_redirect()
                
                # Get document table and document rows
                document_table = self.wait.until(
                    EC.presence_of_element_located(config.DOCUMENT_TABLE_LOCATOR)
                )
                document_rows = document_table.find_elements(
                    config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                    config.DOCUMENT_TABLE_DIV_LOCATOR[1]
                )
                
                if len(document_rows) <= 0:
                    logger.warning("No document rows found in table")
                    break
                
                # Get retry count from config (default: 3)
                download_retry_count = int(self.config.get('DOWNLOAD_RETRY_COUNT', 3))
                
                # If only 1 document, download single with retry
                if client_object.max_total_documents == 1:
                    for retry_attempt in range(download_retry_count):
                        try:
                            self.download_single_document_in_page(document_table, 0, client_object.document_list)
                            break  # Success, exit retry loop
                        except FileDownloadTimeoutError as e:
                            if retry_attempt < download_retry_count - 1:
                                logger.warning(f"File download timeout (attempt {retry_attempt + 1}/{download_retry_count}). Retrying...")
                                time.sleep(10)  # Wait before retry
                            else:
                                logger.error(f"File download timeout after {download_retry_count} attempts")
                                raise
                        except Exception as e:
                            # For other errors, don't retry, just raise
                            raise
                    break  # Exit page loop after single document download
                
                # If multiple documents, try download multiple with retry
                if client_object.max_total_documents > 1:
                    for retry_attempt in range(download_retry_count):
                        try:
                            self.download_multiple_documents_in_page(document_table, client_object.document_list)
                            break  # Success, exit retry loop
                        except FileDownloadTimeoutError as e:
                            if retry_attempt < download_retry_count - 1:
                                logger.warning(f"File download timeout (attempt {retry_attempt + 1}/{download_retry_count}). Retrying...")
                                time.sleep(10)  # Wait before retry
                            else:
                                logger.error(f"File download timeout after {download_retry_count} attempts")
                                raise
                        except Exception as e:
                            # For other errors, don't retry, just raise
                            raise

                # Click next page button to load next page
                if page_num < total_pages:
                    try:
                        next_page_btn = self.wait.until(
                            EC.element_to_be_clickable(config.NEXT_PAGE_BTN_LOCATOR)
                        )
                        btn_class = next_page_btn.get_attribute('class') or ''
                        if 'disabled' not in btn_class:
                            next_page_btn.click()
                            logger.info(f"Switched to page {page_num + 1}")
                            time.sleep(4)
                        else:
                            logger.warning("Next page button is disabled")
                            break
                    except TimeoutException:
                        logger.warning("Next page button not found")
                        break
            
            logger.info(f"Completed export documents for client: {client_object.client_name}")
            
        except (DocumentExportError, WebNavigationError, FileDownloadTimeoutError, FileOperationError, FileNotFoundError, ZipFileError):
            raise
        except Exception as e:
            error_msg = f"Error exporting documents: {str(e)}"
            raise DocumentExportError(error_msg) from e
    
    def download_multiple_documents_in_page(self, document_table, document_list):
        """
        Download multiple documents từ một page
        
        Args:
            document_table: Document table element
            document_list: List các Document objects
            
        Raises:
            DocumentExportError: Nếu có lỗi nghiêm trọng khi download multiple documents
        """
        try:
            logger.info("Downloading multiple documents in page...")
            
            # Select all on current page
            headers_list = self.wait.until(
                EC.presence_of_all_elements_located(config.DOCUMENT_HEADERS_LOCATOR)
            )
            select_all_checkout = headers_list[-1]
            select_all_checkout.click()
            time.sleep(1)
            
            # Click download multiple files
            download_document_btns = self.wait.until(
                EC.presence_of_all_elements_located(config.DOCUMENT_ACTION_BTNS_LOCALTOR)
            )
            download_document_btn = download_document_btns[0]
            download_document_btn.click()
            time.sleep(1)
            
            # Find export document button
            export_document_btns = self.wait.until(
                EC.presence_of_all_elements_located(config.EXPORT_DOCUMENT_BTNS_LOCALTOR)
            )
            export_document_btn = export_document_btns[0]
            
            # Check if button is disabled
            btn_class = export_document_btn.get_attribute('class') or ''
            if 'disabled' in btn_class:
                logger.warning("Export button is disabled, switching to download individual files")
                download_document_btn.click()
                select_all_checkout.click()
                # Download individual files
                document_rows = document_table.find_elements(
                    config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                    config.DOCUMENT_TABLE_DIV_LOCATOR[1]
                )
                for row_index in range(len(document_rows)):
                    try:
                        self.download_single_document_in_page(document_table, row_index, document_list)
                        time.sleep(2) # Wait before downloading next document
                    except DocumentExportError as e:
                        logger.warning(f"Error downloading single document at row {row_index}: {str(e)}")
                return  # Early return since already downloaded individual files
            
            export_document_btn.click()
            time.sleep(1)

            # Click OK button
            btn_ok = self.wait.until(
                EC.presence_of_element_located(config.OK_BTN_LOCALTOR)
            )
            btn_ok.click()
            time.sleep(15)
            select_all_checkout.click()
            
            # Wait for zip file to download
            try:
                zip_file_path = wait_for_file_download(self.download_dir, expected_extension=".zip", timeout=120)
            except FileDownloadTimeoutError:
                logger.warning("Timeout: File not found after 120 seconds")
                return
            except Exception as e:
                raise
            # Move zip to 0_zip_ directory
            zip_dest_path = move_zip_to_storage(zip_file_path, self.zip_dir)
            
            # Create client folder in 0_zip_ (only create once)
            # Format: client_name_client_number_zip
            zip_client_folder_name = f"{document_list[0].client_name}_{document_list[0].client_object.client_number}_zip"
            zip_client_folder_path = os.path.join(self.zip_dir, zip_client_folder_name)
            os.makedirs(zip_client_folder_path, exist_ok=True)
            
            extract_zip(zip_dest_path, zip_client_folder_path)
            
            # Remove zip file after extraction
            remove_file(zip_dest_path)
            
            # Loop through each file in zip (after extraction)
            # Files downloaded this way will have document id at the end
            # Format: name_document_id.ext
            if not os.path.exists(zip_client_folder_path):
                error_msg = f"Zip folder does not exist: {zip_client_folder_path}"
                logger.error(error_msg)
                raise DocumentExportError(error_msg)
            
            # Loop through all files in extracted zip folder
            for root, dirs, files in os.walk(zip_client_folder_path):
                for file_name in files:
                    file_path = os.path.join(root, file_name)
                    
                    # Extract document_id from file name (file format: name_document_id.ext)
                    # Find document_id in file name
                    doc_id = None
                    found_document_object = None
                    for doc in document_list:
                        # File from zip has format: expected_name_document_id.ext
                        # Check if document_id is in file name
                        if doc.document_id in file_name:
                            doc_id = doc.document_id
                            found_document_object = doc
                            break
                    
                    if not doc_id or not found_document_object:
                        logger.warning(f"Document ID or document object not found for doc_id: {doc_id}")
                        continue
                    
                    # Check if file has correct name
                    # Files extracted from zip don't need renaming, if name doesn't match then error
                    expected_file_name = found_document_object.document_name_with_id
                    if file_name.lower() != expected_file_name.lower():
                        error_msg = f"File name mismatch: expected '{expected_file_name}', got '{file_name}'"
                        logger.error(error_msg)
                        found_document_object.set_download_status("Error", error_msg)
                        continue
                    
                    # Move file to document_folder_path
                    if not os.path.exists(found_document_object.document_folder_path):
                        os.makedirs(found_document_object.document_folder_path, exist_ok=True)
                    
                    dest_path = found_document_object.document_file_path
                    move_file(file_path, dest_path)
                    
                    # Update log in Excel
                    download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row_index = self.excel_handler.get_document_row_index(
                        found_document_object.client_name, found_document_object.client_object.client_number, found_document_object.document_id
                    )
                    
                    self.excel_handler.update_document_row(
                        row_index,
                        download_status="Success",
                        download_desc="File downloaded successfully",
                        file_name=found_document_object.document_name_with_id,
                        file_path=dest_path,
                        folder_category=found_document_object.category_name,
                        download_time=download_time
                    )
                    
                    found_document_object.set_download_status("Success", "File downloaded successfully", download_time)
                    logger.info(f"Successfully downloaded document {found_document_object.document_id}")

        except (DocumentExportError, FileDownloadTimeoutError, FileOperationError, FileNotFoundError, ZipFileError):
            raise
        except Exception as e:
            error_msg = f"Error downloading multiple documents: {str(e)}"
            raise DocumentExportError(error_msg) from e

    def download_single_document_in_page(self, document_table, row_index, document_list):
        """
        Download single document from a page
        
        Args:
            document_table: Document table element
            row_index: Index of row to download
            document_list: List of Document objects
            
        Raises:
            DocumentExportError: If there is an error downloading single document
        """
        try:
            logger.info(f"Downloading single document at row {row_index}...")
            
            # Get document rows
            document_rows = document_table.find_elements(
                config.DOCUMENT_TABLE_DIV_LOCATOR[0], 
                config.DOCUMENT_TABLE_DIV_LOCATOR[1]
            )
            
            if row_index >= len(document_rows):
                error_msg = f"Row index {row_index} exceeds number of rows"
                logger.error(error_msg)
                raise DocumentExportError(error_msg)
            
            row = document_rows[row_index]
            
            # Get document data cells
            document_data_cells = row.find_elements(
                config.DOCUMENT_DATA_CELL_LOCATOR[0], 
                config.DOCUMENT_DATA_CELL_LOCATOR[1]
            )
            
            if len(document_data_cells) < 10:
                error_msg = "Not enough cells found in document row"
                logger.error(error_msg)
                raise DocumentExportError(error_msg)
            
            # Get document ID from row
            row_doc_id = document_data_cells[9].text.strip()
            
            # Find corresponding document object
            document = None
            for doc in document_list:
                if doc.document_id == row_doc_id:
                    document = doc
                    break
            
            if not document:
                error_msg = f"Document object not found for doc_id: {row_doc_id}"
                logger.error(error_msg)
                raise DocumentExportError(error_msg)
            
            # Find export button
            document_first_cell = row.find_elements(
                config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[0], 
                config.DOCUMENT_ROW_FIRST_CELL_LOCATOR[1]
            )
            
            if len(document_first_cell) == 0:
                error_msg = "Export button not found"
                logger.error(error_msg)
                raise DocumentExportError(error_msg)
            
            btns = document_first_cell[0].find_elements(By.TAG_NAME, "button")
            if len(btns) < 3:
                error_msg = "Export button not found"
                logger.error(error_msg)
                raise DocumentExportError(error_msg)
            
            btn_export = btns[2]
            btn_export.click()
            logger.info("Clicked export button")
            time.sleep(5)

            # Wait for file to download
            downloaded_file_path = wait_for_file_download(self.download_dir, timeout=120)
            
            # Check if downloaded file name matches document.document_name_without_id
            downloaded_file_name = os.path.basename(downloaded_file_path)
            expected_file_name = document.document_name_without_id
            
            if downloaded_file_name.lower() != expected_file_name.lower():
                error_msg = f"File name mismatch: expected '{expected_file_name}', got '{downloaded_file_name}'"
                logger.error(error_msg)
                document.set_download_status("Error", error_msg)
                # Remove incorrect file to avoid confusion
                remove_file(downloaded_file_path)
                raise DocumentExportError(error_msg)
            
            logger.info(f"Confirmed file has correct name: {downloaded_file_name}")
            
            # File downloaded in single mode so file name will be document_obj.document_name_without_id
            # Must rename to document_obj.document_name_with_id
            renamed_file_path = rename_file_with_doc_id(downloaded_file_path, document.document_id)
            
            # Move file to document_folder_path
            if not os.path.exists(document.document_folder_path):
                os.makedirs(document.document_folder_path, exist_ok=True)
            
            dest_path = document.document_file_path
            move_file(renamed_file_path, dest_path)
            
            # Update log in Excel
            download_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row_index_excel = self.excel_handler.get_document_row_index(
                document.client_name, document.client_object.client_number, document.document_id
            )
            self.excel_handler.update_document_row(
                row_index_excel,
                download_status="Success",
                download_desc="File downloaded successfully",
                file_name=document.document_name_with_id,
                file_path=dest_path,
                folder_category=document.category_name,
                download_time=download_time
            )
            
            document.set_download_status("Success", "File downloaded successfully", download_time)
            logger.info(f"Successfully downloaded document {document.document_id}")
            
        except (DocumentExportError, FileDownloadTimeoutError, FileOperationError, FileNotFoundError):
            raise
        except Exception as e:
            error_msg = f"Error downloading single document: {str(e)}"
            raise DocumentExportError(error_msg) from e


def main():
    """
    Initialize GofileRoomDownloader object, setup driver, login
    From config in .env file, get Excel file path containing client list
    Initialize excel_handler
    Get client list from excel_handler
    Loop through clients with status = Pending
    Each loop:
        - If status is not pending: skip
        - If status is pending:
            - Create client object
            - Perform search_client
            - Perform export_csv_file
            - Perform export_documents
    """
    try:
        # Initialize GofileRoomDownloader
        downloader = GofileRoomDownloader()
        
        # Check if Excel file exists
        if not os.path.exists(downloader.excel_file_path):
            logger.error(f"Excel file does not exist: {downloader.excel_file_path}")
            return
        
        # Load workbook and initialize ExcelHandler
        # Try loading with data_only and keep_links=False to avoid XML parsing issues
        workbook = None
        try:
            workbook = load_workbook(
                downloader.excel_file_path,
                data_only=True,  # Only read values, ignore formulas
                keep_links=False  # Don't keep external links
            )
            logger.info("Successfully loaded workbook with data_only=True")
        except (ValueError, Exception) as e:
            # If fails, try with read_only mode as fallback
            logger.warning(f"Failed to load workbook with data_only=True: {str(e)}")
            logger.warning("Trying to load with read_only=True as fallback...")
            try:
                workbook = load_workbook(
                    downloader.excel_file_path,
                    read_only=True  # Read-only mode, may help with corrupted files
                )
                logger.info("Successfully loaded workbook with read_only=True")
                logger.warning("Note: Workbook is in read-only mode. Saving may not work properly.")
            except Exception as e2:
                error_msg = f"Unable to read workbook: {str(e2)}"
                logger.error(error_msg)
                logger.error("This is most probably because the workbook source files contain some invalid XML.")
                logger.error("Possible causes:")
                logger.error("1. File Excel is corrupted")
                logger.error("2. File Excel has VBA macros or unsupported features")
                logger.error("3. File Excel is currently open in Excel application")
                logger.error("4. File Excel has invalid table structures")
                logger.error("")
                logger.error("Solutions:")
                logger.error("1. Close Excel if it's open")
                logger.error("2. Try opening the file in Excel and save it again (File > Save As > Excel Workbook)")
                logger.error("3. Try creating a new Excel file and copy data manually")
                logger.error("4. Check if file is not corrupted")
                logger.error("5. Remove any VBA macros or unsupported features from the file")
                raise ExcelHandlerError(error_msg) from e2
        
        if workbook is None:
            error_msg = "Failed to load workbook: Unknown error"
            logger.error(error_msg)
            raise ExcelHandlerError(error_msg)
        client_list_sheet_name = downloader.config.get('CLIENT_LIST_SHEET_NAME', 'Client List GFR')
        document_list_sheet_name = downloader.config.get('DOCUMENT_LIST_SHEET_NAME', 'Download Document Log')
        
        if client_list_sheet_name not in workbook.sheetnames:
            logger.error(f"Sheet '{client_list_sheet_name}' does not exist in Excel file")
            return
        
        if document_list_sheet_name not in workbook.sheetnames:
            logger.error(f"Sheet '{document_list_sheet_name}' does not exist in Excel file")
            return
        
        client_list_sheet = workbook[client_list_sheet_name]
        document_list_sheet = workbook[document_list_sheet_name]
        
        # Initialize ExcelHandler
        downloader.excel_handler = ExcelHandler(
            workbook, client_list_sheet, document_list_sheet, downloader.excel_file_path
        )
        
        # Get client list from excel_handler (only get status = "Pending")
        client_list = downloader.excel_handler.get_client_list(status_filter="Pending")
        
        if not client_list:
            logger.info("No clients with status = Pending")
            return
        
        logger.info(f"Found {len(client_list)} clients with status = Pending")
        
        # Loop to process each client
        consecutive_errors = 0
        max_consecutive_errors = int(downloader.config.get('MAX_CONSECUTIVE_ERRORS', 10))
        error_summary = []
        
        for client_info in client_list:
            try:
                logger.info(f"\n{'=' * 80}")
                logger.info(f"Starting to process client: {client_info['client_name']} ({client_info['client_number']})")
                logger.info(f"{'=' * 80}\n")
                
                # Delete all files in download_dir before processing new client
                # (keep subdirectories like 0_csv_, 0_zip_)
                clean_download_dir(downloader.download_dir)
                
                # Create client object
                client_object = Client(client_info['client_name'], client_info['client_number'])
                
                # Perform search_client
                downloader.search_client(client_object)
                
                # Initialize folders after client is successfully found
                client_object.initialize_folders()
                downloader.excel_handler.update_client_row(
                    client_info['row_index'],
                    status="InProgress",
                    description="Downloading documents...",
                    total_documents=client_object.max_total_documents,
                    num_files_downloaded=client_object.get_number_of_downloaded_documents(),
                    client_folder_path=client_object.client_folder_path
                )

                # Check number of documents
                if client_object.max_total_documents == 0:
                    logger.warning(f"Client {client_info['client_name']} has no documents")
                    downloader.excel_handler.update_client_row(
                        client_info['row_index'],
                        status="Success",
                        description="Client has no document",
                        total_documents=0,
                        num_files_downloaded=0,
                        client_folder_path=client_object.client_folder_path
                    )
                    downloader.excel_handler.save_workbook()
                    continue
                
                # Perform export_csv_file
                downloader.export_csv_file(client_object)
                # Perform export_documents
                downloader.export_documents(client_object)
                
                # Update status to success
                if client_object.get_number_of_downloaded_documents() == client_object.max_total_documents:
                    c_status = "Success"
                    c_description = "Download client documents successfully"
                else:
                    c_status = "Warning"
                    c_description = "Download client documents with some errors"
                    
                downloader.excel_handler.update_client_row(
                    client_info['row_index'],
                    status=c_status,
                    description=c_description,
                    total_documents=client_object.max_total_documents,
                    num_files_downloaded=client_object.get_number_of_downloaded_documents(),
                    client_folder_path=client_object.client_folder_path
                )
                downloader.excel_handler.save_workbook()
                
                logger.info(f"Completed processing client: {client_info['client_name']}")
                
                # Reset consecutive errors on success
                consecutive_errors = 0
                error_summary = []
                
            except Exception as e:
                error_msg = f"Error processing client {client_info['client_name']}: {str(e)}"
                logger.error(error_msg)
                logger.error(traceback.format_exc())
                consecutive_errors += 1
                error_summary.append({
                    'client_name': client_info['client_name'],
                    'client_number': client_info['client_number'],
                    'error': error_msg
                })
                
                downloader.excel_handler.update_client_row(
                    client_info['row_index'],
                    status="Error",
                    description=error_msg,
                    total_documents=client_object.max_total_documents,
                    num_files_downloaded=client_object.get_number_of_downloaded_documents(),
                    client_folder_path=client_object.client_folder_path
                )
                downloader.excel_handler.save_workbook()
                
                # Check consecutive errors
                if consecutive_errors >= max_consecutive_errors:
                    logger.critical(f"Found {consecutive_errors} consecutive client errors, stopping automation and sending email")
                    if downloader.email_handler:
                        downloader.email_handler.send_critical_error_email(consecutive_errors, {
                            'total_errors': consecutive_errors,
                            'errors': error_summary[-max_consecutive_errors:]
                        })
                    break
                continue
        
        logger.info(f"\n{'=' * 80}")
        logger.info(f"AUTOMATION COMPLETED")
        logger.info(f"{'=' * 80}\n")
        
    except Exception as e:
        error_msg = f"Critical error running automation: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        
        # Send email when script stops due to critical error
        if 'downloader' in locals() and downloader.email_handler:
            downloader.email_handler.send_error_email(
                subject="[CRITICAL] GOFILEROOM Downloader - Script Stopped Due to Error",
                error_message=error_msg,
                error_details={
                    'Exception Type': type(e).__name__,
                    'Exception Message': str(e),
                    'Traceback': traceback.format_exc()[:1000]  # Limit to 1000 characters
                }
            )
    finally:
        # Cleanup
        if 'downloader' in locals() and downloader and downloader.driver:
            downloader.driver.quit()
            logger.info("Browser closed")


if __name__ == "__main__":
    main()
